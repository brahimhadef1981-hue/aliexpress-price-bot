#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AliExpress Price Monitor Bot - Render.com Compatible Version
With Conflict Handling
"""

import os
import re
import sys
import time
import asyncio
import hashlib
import hmac
import signal
import logging
import traceback

# Setup logging FIRST
logging.basicConfig(
    format='%(asctime)s - %(levelname)s - %(message)s',
    level=logging.INFO,
    stream=sys.stdout
)
logger = logging.getLogger(__name__)

# Print startup info immediately
logger.info("="*60)
logger.info("🚀 STARTING ALIEXPRESS PRICE MONITOR BOT")
logger.info(f"Python version: {sys.version}")
logger.info(f"Working directory: {os.getcwd()}")
logger.info("="*60)

# Check and install missing modules
try:
    import aiohttp
    logger.info("✅ aiohttp imported")
except ImportError as e:
    logger.error(f"❌ Failed to import aiohttp: {e}")
    sys.exit(1)

try:
    import ssl
    import certifi
    logger.info("✅ ssl and certifi imported")
except ImportError as e:
    logger.error(f"❌ Failed to import ssl/certifi: {e}")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    logger.info("✅ openpyxl imported")
except ImportError as e:
    logger.error(f"❌ Failed to import openpyxl: {e}")
    sys.exit(1)

try:
    from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
    from telegram.ext import (
        Application,
        CommandHandler,
        MessageHandler,
        CallbackQueryHandler,
        ConversationHandler,
        ContextTypes,
        filters
    )
    from telegram.error import BadRequest, Conflict, NetworkError, TimedOut
    logger.info("✅ python-telegram-bot imported")
except ImportError as e:
    logger.error(f"❌ Failed to import telegram: {e}")
    logger.error("Try: pip install python-telegram-bot[job-queue]==21.3")
    sys.exit(1)

from datetime import datetime, timedelta
from typing import Dict, Any, Optional, List

# Reduce noise from other loggers
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)
logging.getLogger("telegram").setLevel(logging.WARNING)
logging.getLogger("aiohttp").setLevel(logging.WARNING)

# ============================================================================
# CONFIGURATION
# ============================================================================
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "8354835888:AAF_F1KR40K6nmI_RwkDPwUa74L__CNuY3s")
ALIEXPRESS_APP_KEY = os.getenv("ALIEXPRESS_APP_KEY", "519492")
ALIEXPRESS_APP_SECRET = os.getenv("ALIEXPRESS_APP_SECRET", "R2Zl1pe2p47dFFjXz30546XTwu4JcFlk")
ALIEXPRESS_TRACKING_ID = os.getenv("ALIEXPRESS_TRACKING_ID", "hadef")

# Validate token
if not TELEGRAM_BOT_TOKEN or TELEGRAM_BOT_TOKEN == "YOUR_TOKEN_HERE":
    logger.error("❌ TELEGRAM_BOT_TOKEN is not set!")
    sys.exit(1)

logger.info(f"✅ Bot token loaded: {TELEGRAM_BOT_TOKEN[:10]}...")

# File paths - use /tmp for Render.com or current directory
if os.path.exists("/tmp"):
    DATA_DIR = "/tmp/bot_data"
else:
    DATA_DIR = os.path.dirname(os.path.abspath(__file__))

os.makedirs(DATA_DIR, exist_ok=True)
logger.info(f"📁 Data directory: {DATA_DIR}")

USERS_FILE = os.path.join(DATA_DIR, "users.xlsx")
PRODUCTS_FILE = os.path.join(DATA_DIR, "products.xlsx")
PRICE_HISTORY_FILE = os.path.join(DATA_DIR, "price_history.xlsx")

# Monitoring configuration
CONCURRENT_REQUESTS = 10
REQUEST_DELAY = 1
MONITORING_INTERVAL = 300  # 5 minutes
PRODUCTS_PER_CYCLE = 100
RATE_LIMIT_RETRY_DELAY = 30
MAX_RETRIES = 3
REQUEST_TIMEOUT = 15

# Monthly update configuration
MONTHLY_UPDATE_REMINDER_DAYS = 30
UPDATE_RESPONSE_DEADLINE_DAYS = 3
MONTHLY_CHECK_INTERVAL = 86400

# Conversation states
SELECTING_COUNTRY, ENTERING_LINK, CHANGING_COUNTRY, MANAGING_PRODUCTS, VIEWING_HISTORY = range(5)

# Global API instance
api_instance = None

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def get_ssl_context():
    """Get SSL context for aiohttp"""
    try:
        ssl_context = ssl.create_default_context(cafile=certifi.where())
        return ssl_context
    except Exception as e:
        logger.warning(f"Could not create SSL context with certifi: {e}")
        return ssl.create_default_context()


async def clear_webhook_and_updates():
    """Clear any existing webhook and pending updates before starting"""
    logger.info("🔄 Clearing webhook and pending updates...")
    
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/deleteWebhook?drop_pending_updates=true"
    
    try:
        ssl_context = get_ssl_context()
        connector = aiohttp.TCPConnector(ssl=ssl_context)
        
        async with aiohttp.ClientSession(connector=connector) as session:
            async with session.get(url, timeout=aiohttp.ClientTimeout(total=10)) as response:
                result = await response.json()
                if result.get('ok'):
                    logger.info("✅ Webhook cleared successfully")
                else:
                    logger.warning(f"⚠️ Webhook clear response: {result}")
                return result.get('ok', False)
    except Exception as e:
        logger.warning(f"⚠️ Could not clear webhook: {e}")
        return False


async def get_bot_info():
    """Get bot info to verify token is valid"""
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getMe"
    
    try:
        ssl_context = get_ssl_context()
        connector = aiohttp.TCPConnector(ssl=ssl_context)
        
        async with aiohttp.ClientSession(connector=connector) as session:
            async with session.get(url, timeout=aiohttp.ClientTimeout(total=10)) as response:
                result = await response.json()
                if result.get('ok'):
                    bot_info = result.get('result', {})
                    logger.info(f"✅ Bot verified: @{bot_info.get('username', 'unknown')}")
                    return True
                else:
                    logger.error(f"❌ Invalid bot token: {result}")
                    return False
    except Exception as e:
        logger.error(f"❌ Could not verify bot: {e}")
        return False


# ============================================================================
# EXCEL MANAGEMENT
# ============================================================================
class ExcelManager:
    @staticmethod
    def init_excel_files():
        """Initialize Excel files"""
        try:
            if not os.path.exists(USERS_FILE):
                wb = Workbook()
                ws = wb.active
                ws.title = "Users"
                ws.append(["User ID", "Username", "Country", "Date Added", "Last Update Reminder", 
                          "Update Deadline", "Needs Update Response"])
                wb.save(USERS_FILE)
                logger.info(f"✅ Created {USERS_FILE}")

            if not os.path.exists(PRODUCTS_FILE):
                wb = Workbook()
                ws = wb.active
                ws.title = "Products"
                ws.append(["User ID", "Product ID", "Product URL", "Title", "Current Price", 
                          "Original Price", "Currency", "Image URL", "Country", "Date Added", "Last Checked"])
                wb.save(PRODUCTS_FILE)
                logger.info(f"✅ Created {PRODUCTS_FILE}")
            
            if not os.path.exists(PRICE_HISTORY_FILE):
                wb = Workbook()
                ws = wb.active
                ws.title = "Price History"
                ws.append(["User ID", "Product ID", "Product Title", "Old Price", "New Price", 
                          "Change Amount", "Change Percent", "Currency", "Date"])
                wb.save(PRICE_HISTORY_FILE)
                logger.info(f"✅ Created {PRICE_HISTORY_FILE}")
            
            logger.info("✅ All Excel files initialized")
                
        except Exception as e:
            logger.error(f"❌ Error creating Excel files: {e}")
            logger.error(traceback.format_exc())
            raise

    @staticmethod
    def save_user(user_id: int, username: str, country: str):
        try:
            wb = load_workbook(USERS_FILE)
            ws = wb.active
            
            user_exists = False
            for row in ws.iter_rows(min_row=2, max_col=7):
                if row[0].value == user_id:
                    row[2].value = country
                    user_exists = True
                    break
            
            if not user_exists:
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ws.append([user_id, username, country, now, "", "", "No"])
            
            wb.save(USERS_FILE)
        except Exception as e:
            logger.error(f"Error saving user: {e}")

    @staticmethod
    def get_user_country(user_id: int) -> Optional[str]:
        try:
            if not os.path.exists(USERS_FILE):
                return None
            wb = load_workbook(USERS_FILE)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, max_col=3):
                if row[0].value == user_id:
                    return row[2].value
        except Exception as e:
            logger.error(f"Error getting user country: {e}")
        return None

    @staticmethod
    def update_user_products_country(user_id: int, new_country: str):
        try:
            if not os.path.exists(PRODUCTS_FILE):
                return 0
            wb = load_workbook(PRODUCTS_FILE)
            ws = wb.active
            
            updated_count = 0
            for row in ws.iter_rows(min_row=2):
                if row[0].value == user_id:
                    row[8].value = new_country
                    updated_count += 1
            
            wb.save(PRODUCTS_FILE)
            return updated_count
        except Exception as e:
            logger.error(f"Error updating products country: {e}")
            return 0

    @staticmethod
    def save_product(user_id: int, product_id: str, product_url: str, title: str, 
                    price: float, original_price: float, currency: str, image_url: str, country: str):
        try:
            wb = load_workbook(PRODUCTS_FILE)
            ws = wb.active
            
            product_exists = False
            for row in ws.iter_rows(min_row=2):
                if row[0].value == user_id and row[1].value == product_id:
                    row[2].value = product_url
                    row[4].value = price
                    row[5].value = original_price
                    row[8].value = country
                    row[10].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    product_exists = True
                    break
            
            if not product_exists:
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ws.append([user_id, product_id, product_url, title, price, original_price, 
                          currency, image_url, country, now, now])
            
            wb.save(PRODUCTS_FILE)
            logger.info(f"✅ Product {product_id} saved for user {user_id}")
        except Exception as e:
            logger.error(f"Error saving product: {e}")

    @staticmethod
    def get_all_products() -> List[Dict]:
        products = []
        try:
            if not os.path.exists(PRODUCTS_FILE):
                return products
            wb = load_workbook(PRODUCTS_FILE)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    products.append({
                        'user_id': row[0],
                        'product_id': row[1],
                        'product_url': row[2],
                        'title': row[3],
                        'current_price': row[4] or 0,
                        'original_price': row[5] or 0,
                        'currency': row[6] or 'USD',
                        'image_url': row[7],
                        'country': row[8] or 'US',
                        'date_added': row[9],
                        'last_checked': row[10]
                    })
        except Exception as e:
            logger.error(f"Error getting products: {e}")
        
        return products

    @staticmethod
    def get_products_to_check(limit: int) -> List[Dict]:
        all_products = ExcelManager.get_all_products()
        
        if not all_products:
            return []
        
        def get_last_checked_time(product):
            last_checked = product.get('last_checked')
            if not last_checked or last_checked == 'Never' or last_checked == '':
                return datetime.min
            try:
                return datetime.strptime(str(last_checked), "%Y-%m-%d %H:%M:%S")
            except:
                return datetime.min
        
        sorted_products = sorted(all_products, key=get_last_checked_time)
        return sorted_products[:limit]

    @staticmethod
    def update_product_price(user_id: int, product_id: str, new_price: float, country: str = None, product_url: str = None):
        try:
            wb = load_workbook(PRODUCTS_FILE)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2):
                if row[0].value == user_id and row[1].value == product_id:
                    if product_url:
                        row[2].value = product_url
                    row[4].value = new_price
                    if country:
                        row[8].value = country
                    row[10].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    break
            
            wb.save(PRODUCTS_FILE)
        except Exception as e:
            logger.error(f"Error updating product price: {e}")

    @staticmethod
    def save_price_change(user_id: int, product_id: str, title: str, old_price: float, 
                         new_price: float, currency: str):
        try:
            if abs(new_price - old_price) < 0.01:
                return
            
            wb = load_workbook(PRICE_HISTORY_FILE)
            ws = wb.active
            
            change = new_price - old_price
            change_percent = ((new_price - old_price) / old_price * 100) if old_price > 0 else 0
            
            ws.append([
                user_id,
                product_id,
                title,
                old_price,
                new_price,
                round(change, 2),
                round(change_percent, 2),
                currency,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ])
            
            wb.save(PRICE_HISTORY_FILE)
            logger.info(f"💰 Price change: {title[:30]} ${old_price:.2f} → ${new_price:.2f}")
        except Exception as e:
            logger.error(f"Error saving price change: {e}")

    @staticmethod
    def get_user_products(user_id: int) -> List[Dict]:
        all_products = ExcelManager.get_all_products()
        return [p for p in all_products if p['user_id'] == user_id]

    @staticmethod
    def delete_product(user_id: int, product_id: str) -> bool:
        try:
            wb = load_workbook(PRODUCTS_FILE)
            ws = wb.active
            
            rows_to_delete = []
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row[0].value == user_id and row[1].value == product_id:
                    rows_to_delete.append(idx)
            
            for row_idx in reversed(rows_to_delete):
                ws.delete_rows(row_idx)
            
            wb.save(PRODUCTS_FILE)
            return True
        except Exception as e:
            logger.error(f"Error deleting product: {e}")
            return False

    @staticmethod
    def get_price_history(user_id: int, product_id: str, months: int = None) -> List[Dict]:
        history = []
        try:
            if not os.path.exists(PRICE_HISTORY_FILE):
                return history
            wb = load_workbook(PRICE_HISTORY_FILE)
            ws = wb.active
            
            cutoff_date = None
            if months:
                cutoff_date = datetime.now() - timedelta(days=months * 30)
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == user_id and row[1] == product_id:
                    try:
                        record_date = datetime.strptime(str(row[8]), "%Y-%m-%d %H:%M:%S")
                        
                        if cutoff_date and record_date < cutoff_date:
                            continue
                        
                        history.append({
                            'title': row[2],
                            'old_price': row[3],
                            'new_price': row[4],
                            'change_amount': row[5],
                            'change_percent': row[6],
                            'currency': row[7],
                            'date': row[8]
                        })
                    except:
                        pass
        except Exception as e:
            logger.error(f"Error getting price history: {e}")
        
        return sorted(history, key=lambda x: x['date'], reverse=True)

    @staticmethod
    def get_all_user_price_history(user_id: int, months: int = None) -> Dict[str, List[Dict]]:
        products = ExcelManager.get_user_products(user_id)
        history_by_product = {}
        
        for product in products:
            history = ExcelManager.get_price_history(user_id, product['product_id'], months)
            if history:
                history_by_product[product['product_id']] = {
                    'product': product,
                    'history': history
                }
        
        return history_by_product

    @staticmethod
    def set_update_reminder(user_id: int):
        try:
            wb = load_workbook(USERS_FILE)
            ws = wb.active
            
            now = datetime.now()
            deadline = now + timedelta(days=UPDATE_RESPONSE_DEADLINE_DAYS)
            
            for row in ws.iter_rows(min_row=2):
                if row[0].value == user_id:
                    row[4].value = now.strftime("%Y-%m-%d %H:%M:%S")
                    row[5].value = deadline.strftime("%Y-%m-%d %H:%M:%S")
                    row[6].value = "Yes"
                    break
            
            wb.save(USERS_FILE)
        except Exception as e:
            logger.error(f"Error setting update reminder: {e}")

    @staticmethod
    def clear_update_reminder(user_id: int):
        try:
            wb = load_workbook(USERS_FILE)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2):
                if row[0].value == user_id:
                    row[4].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    row[5].value = ""
                    row[6].value = "No"
                    break
            
            wb.save(USERS_FILE)
        except Exception as e:
            logger.error(f"Error clearing update reminder: {e}")

    @staticmethod
    def get_users_needing_reminder() -> List[int]:
        users = []
        try:
            if not os.path.exists(USERS_FILE):
                return users
            wb = load_workbook(USERS_FILE)
            ws = wb.active
            
            now = datetime.now()
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                
                user_id = row[0]
                last_reminder = row[4] if len(row) > 4 else None
                
                if not last_reminder or last_reminder == "":
                    users.append(user_id)
                else:
                    try:
                        last_reminder_date = datetime.strptime(str(last_reminder), "%Y-%m-%d %H:%M:%S")
                        days_since = (now - last_reminder_date).days
                        
                        if days_since >= MONTHLY_UPDATE_REMINDER_DAYS:
                            users.append(user_id)
                    except:
                        users.append(user_id)
        
        except Exception as e:
            logger.error(f"Error getting users needing reminder: {e}")
        
        return users

    @staticmethod
    def get_users_past_deadline() -> List[int]:
        users = []
        try:
            if not os.path.exists(USERS_FILE):
                return users
            wb = load_workbook(USERS_FILE)
            ws = wb.active
            
            now = datetime.now()
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                
                user_id = row[0]
                needs_response = row[6] if len(row) > 6 else None
                deadline = row[5] if len(row) > 5 else None
                
                if needs_response == "Yes" and deadline:
                    try:
                        deadline_date = datetime.strptime(str(deadline), "%Y-%m-%d %H:%M:%S")
                        
                        if now > deadline_date:
                            users.append(user_id)
                    except:
                        pass
        
        except Exception as e:
            logger.error(f"Error getting users past deadline: {e}")
        
        return users

    @staticmethod
    def delete_all_user_data(user_id: int):
        try:
            if os.path.exists(PRODUCTS_FILE):
                wb = load_workbook(PRODUCTS_FILE)
                ws = wb.active
                
                rows_to_delete = []
                for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    if row[0].value == user_id:
                        rows_to_delete.append(idx)
                
                for row_idx in reversed(rows_to_delete):
                    ws.delete_rows(row_idx)
                
                wb.save(PRODUCTS_FILE)
            
            if os.path.exists(PRICE_HISTORY_FILE):
                wb = load_workbook(PRICE_HISTORY_FILE)
                ws = wb.active
                
                rows_to_delete = []
                for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    if row[0].value == user_id:
                        rows_to_delete.append(idx)
                
                for row_idx in reversed(rows_to_delete):
                    ws.delete_rows(row_idx)
                
                wb.save(PRICE_HISTORY_FILE)
            
            logger.info(f"Deleted all data for user {user_id}")
            return True
            
        except Exception as e:
            logger.error(f"Error deleting user data: {e}")
            return False


# ============================================================================
# ALIEXPRESS API CLIENT
# ============================================================================
class AliExpressAPI:
    def __init__(self, app_key: str, app_secret: str, tracking_id: str):
        self.app_key = app_key
        self.app_secret = app_secret
        self.tracking_id = tracking_id
        self.api_url = "https://api-sg.aliexpress.com/sync"
        self.timeout = aiohttp.ClientTimeout(total=REQUEST_TIMEOUT)
        self.session = None
        self._session_lock = asyncio.Lock()

    async def get_session(self):
        async with self._session_lock:
            if self.session is None or self.session.closed:
                ssl_context = get_ssl_context()
                connector = aiohttp.TCPConnector(
                    limit=50,
                    limit_per_host=10,
                    ttl_dns_cache=300,
                    ssl=ssl_context
                )
                self.session = aiohttp.ClientSession(
                    connector=connector,
                    timeout=self.timeout,
                    headers={
                        'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36'
                    }
                )
            return self.session

    async def close_session(self):
        async with self._session_lock:
            if self.session and not self.session.closed:
                await self.session.close()
                await asyncio.sleep(0.25)
                self.session = None

    @staticmethod
    def _now_ms() -> str:
        return str(int(time.time() * 1000))

    def generate_signature(self, params: Dict[str, Any]) -> str:
        params_to_sign = {
            k: str(v) for k, v in params.items()
            if k != "sign" and v is not None and v != ""
        }
        
        sorted_items = sorted(params_to_sign.items(), key=lambda x: x[0])
        canonical = "".join(f"{k}{v}" for k, v in sorted_items)
        
        signature = hmac.new(
            self.app_secret.encode("utf-8"),
            canonical.encode("utf-8"),
            hashlib.md5,
        ).hexdigest().upper()
        
        return signature

    @staticmethod
    def extract_product_id(url: str) -> Optional[str]:
        patterns = [
            r"/item/(\d+)\.html",
            r"/i/(\d+)\.html",
            r"/(\d+)\.html",
            r"item/(\d+)",
            r"/goods/(\d+)",
            r"product/(\d+)",
            r"/dp/(\d+)",
        ]
        for p in patterns:
            m = re.search(p, url)
            if m:
                return m.group(1)
        return None

    @staticmethod
    def build_product_url(product_id: str) -> str:
        return f"https://www.aliexpress.com/item/{product_id}.html"

    async def resolve_shortened_url(self, url: str, max_retries: int = 3) -> str:
        headers = {
            'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        }
        
        session = await self.get_session()
        
        for attempt in range(max_retries):
            try:
                async with session.head(url, allow_redirects=True, headers=headers, timeout=aiohttp.ClientTimeout(total=10)) as response:
                    final_url = str(response.url)
                    
                    if self.extract_product_id(final_url):
                        return final_url
                
            except Exception as e:
                if attempt < max_retries - 1:
                    await asyncio.sleep(2)
                    continue
                return url
        
        return url

    @staticmethod
    def is_shortened_url(url: str) -> bool:
        patterns = [
            "s.click.aliexpress.com",
            "a.aliexpress.com",
            "/e/_",
            "ali.ski",
        ]
        return any(pattern in url.lower() for pattern in patterns)

    @staticmethod
    def is_rate_limited(error_msg: str) -> bool:
        rate_limit_patterns = [
            "frequency exceeds the limit",
            "rate limit",
            "too many requests",
        ]
        return any(pattern in error_msg.lower() for pattern in rate_limit_patterns)

    async def get_product_details(self, product_id: str, country: str = "US", retry_count: int = 0) -> Dict[str, Any]:
        start_time = time.time()
        
        method = "aliexpress.affiliate.productdetail.get"
        
        params = {
            "app_key": self.app_key,
            "format": "json",
            "method": method,
            "sign_method": "hmac",
            "timestamp": self._now_ms(),
            "v": "2.0",
            "tracking_id": self.tracking_id,
            "product_ids": str(product_id),
            "target_currency": "USD",
            "target_language": "EN",
            "country": country,
        }
        
        params["sign"] = self.generate_signature(params)
        
        session = await self.get_session()
        
        try:
            async with session.get(self.api_url, params=params) as response:
                response.raise_for_status()
                data = await response.json()
                
                elapsed_time = time.time() - start_time
                
                if "error_response" in data:
                    error_msg = data["error_response"].get("msg", "API Error")
                    
                    if self.is_rate_limited(error_msg) and retry_count < MAX_RETRIES:
                        wait_time = RATE_LIMIT_RETRY_DELAY * (2 ** retry_count)
                        await asyncio.sleep(wait_time)
                        return await self.get_product_details(product_id, country, retry_count + 1)
                    
                    return {"success": False, "error": error_msg, "time_taken": elapsed_time}
                
                resp_key = None
                for k in data.keys():
                    if k.endswith("_response"):
                        resp_key = k
                        break
                
                if not resp_key:
                    return {"success": False, "error": "Invalid response", "time_taken": elapsed_time}
                
                resp_data = data[resp_key]
                result = resp_data.get("resp_result", {}).get("result", {})
                products = result.get("products", {}).get("product", [])
                
                if not products:
                    return {"success": False, "error": "Product not found", "time_taken": elapsed_time}
                
                product = products[0] if isinstance(products, list) else products
                
                sale_price = product.get("target_sale_price") or product.get("sale_price")
                original_price = product.get("target_original_price") or product.get("original_price")
                
                def to_float(val):
                    if val is None:
                        return None
                    try:
                        price_str = str(val).replace("USD", "").replace("$", "").replace(",", "").strip()
                        return float(price_str)
                    except:
                        return None
                
                current_price = to_float(sale_price)
                orig_price = to_float(original_price)
                
                if current_price is None:
                    return {"success": False, "error": "No price available", "time_taken": elapsed_time}
                
                product_url = self.build_product_url(str(product.get("product_id", product_id)))
                
                return {
                    "success": True,
                    "product_id": str(product.get("product_id", product_id)),
                    "title": product.get("product_title", "N/A"),
                    "price": current_price,
                    "original_price": orig_price or current_price,
                    "currency": "USD",
                    "image_url": product.get("product_main_image_url", ""),
                    "product_url": product_url,
                    "time_taken": elapsed_time
                }
                
        except asyncio.TimeoutError:
            elapsed_time = time.time() - start_time
            return {"success": False, "error": "Request timeout", "time_taken": elapsed_time}
        except Exception as e:
            elapsed_time = time.time() - start_time
            return {"success": False, "error": str(e), "time_taken": elapsed_time}

    async def generate_affiliate_link(self, product_url: str, country: str = "US") -> Optional[str]:
        method = "aliexpress.affiliate.link.generate"
        
        params = {
            "app_key": self.app_key,
            "format": "json",
            "method": method,
            "sign_method": "hmac",
            "timestamp": self._now_ms(),
            "v": "2.0",
            "tracking_id": self.tracking_id,
            "promotion_link_type": "0",
            "source_values": product_url,
        }
        
        params["sign"] = self.generate_signature(params)
        
        session = await self.get_session()
        
        try:
            async with session.post(self.api_url, data=params) as response:
                data = await response.json()
                
                if "error_response" in data:
                    return product_url
                
                resp_data = data.get("aliexpress_affiliate_link_generate_response", {})
                result = resp_data.get("resp_result", {})
                
                if result.get("resp_code") == 200:
                    links = result.get("result", {}).get("promotion_links", {}).get("promotion_link", [])
                    if links:
                        return links[0].get("promotion_link")
                
                return product_url
                
        except Exception as e:
            return product_url


async def get_api_instance():
    global api_instance
    if api_instance is None:
        api_instance = AliExpressAPI(ALIEXPRESS_APP_KEY, ALIEXPRESS_APP_SECRET, ALIEXPRESS_TRACKING_ID)
    return api_instance


# ============================================================================
# MESSAGE HELPERS
# ============================================================================
async def safe_edit_message(query, text, reply_markup=None, parse_mode='HTML'):
    try:
        await query.edit_message_text(text, reply_markup=reply_markup, parse_mode=parse_mode)
    except BadRequest:
        try:
            await query.message.delete()
        except:
            pass
        try:
            await query.message.reply_text(text, reply_markup=reply_markup, parse_mode=parse_mode)
        except:
            pass
    except Exception as e:
        logger.error(f"Error editing message: {e}")
        try:
            await query.message.reply_text(text, reply_markup=reply_markup, parse_mode=parse_mode)
        except:
            pass


# ============================================================================
# TELEGRAM BOT HANDLERS
# ============================================================================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    username = user.username or user.first_name
    logger.info(f"User {user.id} ({username}) started the bot")

    keyboard = [
        [
            InlineKeyboardButton("🇫🇷 France", callback_data="country_FR"),
            InlineKeyboardButton("🇮🇹 Italy", callback_data="country_IT"),
        ],
        [
            InlineKeyboardButton("🇺🇸 United States", callback_data="country_US"),
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    welcome_text = (
        f"👋 <b>Welcome {username}!</b>\n\n"
        "🛍️ <b>AliExpress Price Monitor Bot</b>\n"
        "━━━━━━━━━━━━━━━━━━━━━\n\n"
        "I will help you track AliExpress product prices and notify you of changes!\n\n"
        "📍 <b>Please select your country:</b>"
    )

    await update.message.reply_text(welcome_text, reply_markup=reply_markup, parse_mode='HTML')
    return SELECTING_COUNTRY


async def country_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    country = query.data.split("_")[1]
    user = update.effective_user
    user_id = user.id
    username = user.username or user.first_name

    ExcelManager.save_user(user_id, username, country)
    updated_count = ExcelManager.update_user_products_country(user_id, country)

    context.user_data['country'] = country

    country_flags = {"FR": "🇫🇷", "IT": "🇮🇹", "US": "🇺🇸"}

    message = f"✅ <b>Country Selected: {country_flags.get(country, '')} {country}</b>\n\n"

    if updated_count > 0:
        message += f"🔄 Updated {updated_count} existing products\n\n"

    message += (
        "📎 <b>Now send me an AliExpress product link:</b>\n\n"
        "<i>Supported formats:</i>\n"
        "• <code>https://www.aliexpress.com/item/xxxxx.html</code>\n"
        "• <code>https://s.click.aliexpress.com/e/_xxxxx</code>\n\n"
        "💡 Or use the menu below:"
    )

    keyboard = [
        [InlineKeyboardButton("➕ Add Product", callback_data="add_product")],
        [InlineKeyboardButton("📋 My Products", callback_data="view_myproducts")],
        [InlineKeyboardButton("📊 Price History", callback_data="view_history")],
        [InlineKeyboardButton("ℹ️ Help", callback_data="show_help")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await safe_edit_message(query, message, reply_markup)
    return ENTERING_LINK


async def add_product_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    message = (
        "📎 <b>Send me an AliExpress product link:</b>\n\n"
        "<i>Supported formats:</i>\n"
        "• <code>https://www.aliexpress.com/item/xxxxx.html</code>\n"
        "• <code>https://s.click.aliexpress.com/e/_xxxxx</code>\n\n"
        "💡 Just paste the link and send it to me!"
    )
    
    keyboard = [
        [InlineKeyboardButton("📋 My Products", callback_data="view_myproducts")],
        [InlineKeyboardButton("🔙 Back to Menu", callback_data="back_to_menu")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await safe_edit_message(query, message, reply_markup)
    return ENTERING_LINK


async def handle_link(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.text:
        return ENTERING_LINK
    
    user_id = update.effective_user.id
    product_url = update.message.text.strip()
    
    logger.info(f"User {user_id} sent link: {product_url[:50]}...")

    if "aliexpress" not in product_url.lower():
        keyboard = [
            [InlineKeyboardButton("➕ Add Product", callback_data="add_product")],
            [InlineKeyboardButton("📋 My Products", callback_data="view_myproducts")],
            [InlineKeyboardButton("🔙 Back to Menu", callback_data="back_to_menu")],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(
            "❌ <b>Invalid Link</b>\n\nPlease send a valid AliExpress product link.",
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
        return ENTERING_LINK

    processing_msg = await update.message.reply_text("⏳ Processing...")

    country = ExcelManager.get_user_country(user_id) or "US"
    api = await get_api_instance()

    if api.is_shortened_url(product_url):
        await processing_msg.edit_text("🔗 Resolving shortened URL...")
        product_url = await api.resolve_shortened_url(product_url)

    product_id = api.extract_product_id(product_url)

    if not product_id:
        keyboard = [
            [InlineKeyboardButton("➕ Try Again", callback_data="add_product")],
            [InlineKeyboardButton("🔙 Back to Menu", callback_data="back_to_menu")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await processing_msg.edit_text(
            "❌ <b>Could not extract product ID</b>\n\nPlease send a valid AliExpress link.",
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
        return ENTERING_LINK

    await processing_msg.edit_text("📊 Fetching product details...")

    result = await api.get_product_details(product_id, country)

    if not result.get("success"):
        keyboard = [
            [InlineKeyboardButton("➕ Try Another", callback_data="add_product")],
            [InlineKeyboardButton("🔙 Back to Menu", callback_data="back_to_menu")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await processing_msg.edit_text(
            f"❌ <b>Cannot monitor this product</b>\n\n"
            f"<b>Reason:</b> {result.get('error')}\n\n"
            "💡 Try another product or check back later.",
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
        return ENTERING_LINK

    affiliate_link = await api.generate_affiliate_link(result['product_url'], country)

    ExcelManager.save_product(
        user_id=user_id,
        product_id=product_id,
        product_url=affiliate_link,
        title=result['title'],
        price=result['price'],
        original_price=result['original_price'],
        currency=result['currency'],
        image_url=result['image_url'],
        country=country
    )

    discount = result['original_price'] - result['price']
    discount_percent = (discount / result['original_price'] * 100) if result['original_price'] > 0 else 0

    message = (
        "✅ <b>Product Added Successfully!</b>\n"
        "━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"📦 <b>{result['title'][:80]}...</b>\n\n"
        f"💵 <b>Current:</b> ${result['price']:.2f}\n"
        f"💰 <b>Original:</b> ${result['original_price']:.2f}\n"
    )

    if discount > 0:
        message += f"🏷️ <b>Discount:</b> ${discount:.2f} ({discount_percent:.1f}% OFF)\n"

    message += (
        f"🌍 <b>Country:</b> {country}\n"
        f"🆔 <b>ID:</b> {product_id}\n\n"
        f"🔔 <b>Price monitoring active!</b>\n\n"
        f"🔗 <code>{affiliate_link}</code>"
    )

    keyboard = [
        [InlineKeyboardButton("➕ Add Another Product", callback_data="add_product")],
        [InlineKeyboardButton("📋 My Products", callback_data="view_myproducts")],
        [InlineKeyboardButton("📊 Price History", callback_data="view_history")],
        [InlineKeyboardButton("🔙 Back to Menu", callback_data="back_to_menu")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    if result.get('image_url'):
        try:
            await update.message.reply_photo(
                photo=result['image_url'],
                caption=message,
                reply_markup=reply_markup,
                parse_mode='HTML'
            )
            await processing_msg.delete()
        except Exception as e:
            logger.warning(f"Could not send photo: {e}")
            await processing_msg.edit_text(message, reply_markup=reply_markup, parse_mode='HTML')
    else:
        await processing_msg.edit_text(message, reply_markup=reply_markup, parse_mode='HTML')

    return ENTERING_LINK


async def view_my_products(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if query:
        await query.answer()
    
    user_id = update.effective_user.id
    products = ExcelManager.get_user_products(user_id)

    if not products:
        keyboard = [
            [InlineKeyboardButton("➕ Add Product", callback_data="add_product")],
            [InlineKeyboardButton("🔙 Back", callback_data="back_to_menu")],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        message = (
            "📭 <b>No Products Yet</b>\n\n"
            "You haven't added any products to monitor.\n\n"
            "Click 'Add Product' to start tracking prices!"
        )
        
        if query:
            await safe_edit_message(query, message, reply_markup)
        else:
            await update.message.reply_text(message, reply_markup=reply_markup, parse_mode='HTML')
        return

    message = f"📦 <b>Your Monitored Products ({len(products)}):</b>\n━━━━━━━━━━━━━━━━━━━━━\n\n"

    for i, product in enumerate(products[:5], 1):
        title = product['title'][:40] + "..." if len(str(product['title'])) > 40 else product['title']
        price = product['current_price'] or 0
        message += (
            f"{i}. <b>{title}</b>\n"
            f"   💵 ${price:.2f}\n"
            f"   🌍 {product['country']}\n\n"
        )

    if len(products) > 5:
        message += f"<i>...and {len(products) - 5} more</i>\n\n"

    keyboard = [
        [InlineKeyboardButton("➕ Add Product", callback_data="add_product")],
        [InlineKeyboardButton("🗑️ Manage Products", callback_data="manage_products")],
        [InlineKeyboardButton("📊 Price History", callback_data="view_history")],
        [InlineKeyboardButton("🔙 Back to Menu", callback_data="back_to_menu")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    if query:
        await safe_edit_message(query, message, reply_markup)
    else:
        await update.message.reply_text(message, reply_markup=reply_markup, parse_mode='HTML')


async def manage_products(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = update.effective_user.id
    products = ExcelManager.get_user_products(user_id)

    message = f"🗑️ <b>Manage Products ({len(products)}):</b>\n━━━━━━━━━━━━━━━━━━━━━\n\nSelect a product to delete:\n\n"

    keyboard = []
    for product in products[:10]:
        title = product['title'][:25] + "..." if len(str(product['title'])) > 25 else product['title']
        price = product['current_price'] or 0
        keyboard.append([
            InlineKeyboardButton(
                f"❌ {title} - ${price:.2f}",
                callback_data=f"delete_{product['product_id']}"
            )
        ])

    keyboard.append([InlineKeyboardButton("➕ Add Product", callback_data="add_product")])
    keyboard.append([InlineKeyboardButton("🔙 Back", callback_data="view_myproducts")])
    reply_markup = InlineKeyboardMarkup(keyboard)

    await safe_edit_message(query, message, reply_markup)


async def delete_product_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    product_id = query.data.split("_", 1)[1]
    user_id = update.effective_user.id

    products = ExcelManager.get_user_products(user_id)
    product_title = "Product"
    for p in products:
        if p['product_id'] == product_id:
            product_title = str(p['title'])[:50]
            break

    success = ExcelManager.delete_product(user_id, product_id)

    if success:
        keyboard = [
            [InlineKeyboardButton("➕ Add Product", callback_data="add_product")],
            [InlineKeyboardButton("📋 My Products", callback_data="view_myproducts")],
            [InlineKeyboardButton("🔙 Back to Menu", callback_data="back_to_menu")],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await safe_edit_message(
            query,
            f"✅ <b>Product Deleted</b>\n\n<b>{product_title}</b>\n\nThis product has been removed from monitoring.",
            reply_markup
        )
    else:
        await safe_edit_message(query, "❌ Error deleting product. Please try again.")


async def view_price_history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if query:
        await query.answer()
    
    message = (
        "📊 <b>Price History</b>\n"
        "━━━━━━━━━━━━━━━━━━━━━\n\n"
        "Select time period to view price changes:\n\n"
        "💡 Only actual price changes are archived"
    )

    keyboard = [
        [
            InlineKeyboardButton("1 Month", callback_data="history_1"),
            InlineKeyboardButton("2 Months", callback_data="history_2"),
            InlineKeyboardButton("3 Months", callback_data="history_3"),
        ],
        [
            InlineKeyboardButton("4 Months", callback_data="history_4"),
            InlineKeyboardButton("5 Months", callback_data="history_5"),
            InlineKeyboardButton("6 Months", callback_data="history_6"),
        ],
        [InlineKeyboardButton("📅 All Time", callback_data="history_all")],
        [InlineKeyboardButton("➕ Add Product", callback_data="add_product")],
        [InlineKeyboardButton("🔙 Back", callback_data="view_myproducts")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    if query:
        await safe_edit_message(query, message, reply_markup)
    else:
        await update.message.reply_text(message, reply_markup=reply_markup, parse_mode='HTML')


async def show_price_history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = update.effective_user.id
    period = query.data.split("_")[1]
    
    months = None if period == "all" else int(period)
    history_data = ExcelManager.get_all_user_price_history(user_id, months)

    if not history_data:
        keyboard = [
            [InlineKeyboardButton("➕ Add Products", callback_data="add_product")],
            [InlineKeyboardButton("🔙 Back", callback_data="view_history")],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        period_text = f"last {period} month(s)" if period != "all" else "all time"
        await safe_edit_message(
            query,
            f"📊 <b>No Price Changes</b>\n\nNo price changes recorded for {period_text}.",
            reply_markup
        )
        return

    period_text = f"Last {period} Month(s)" if period != "all" else "All Time"
    message = f"📊 <b>Price History - {period_text}</b>\n━━━━━━━━━━━━━━━━━━━━━\n\n"

    total_changes = 0
    for product_id, data in list(history_data.items())[:5]:
        product = data['product']
        history = data['history']
        total_changes += len(history)
        
        title = str(product['title'])[:40] + "..." if len(str(product['title'])) > 40 else product['title']
        message += f"📦 <b>{title}</b>\n"
        
        for change in history[:3]:
            try:
                date = datetime.strptime(str(change['date']), "%Y-%m-%d %H:%M:%S").strftime("%m/%d")
                emoji = "📉" if float(change['change_amount']) < 0 else "📈"
                message += (
                    f"   {emoji} ${float(change['old_price']):.2f} → ${float(change['new_price']):.2f} "
                    f"({float(change['change_percent']):+.1f}%) - {date}\n"
                )
            except:
                pass
        
        if len(history) > 3:
            message += f"   <i>...and {len(history) - 3} more changes</i>\n"
        
        message += "\n"

    message += f"📈 <b>Total Changes:</b> {total_changes}"

    keyboard = [
        [InlineKeyboardButton("➕ Add Product", callback_data="add_product")],
        [InlineKeyboardButton("🔄 Change Period", callback_data="view_history")],
        [InlineKeyboardButton("🔙 Back to Menu", callback_data="back_to_menu")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await safe_edit_message(query, message, reply_markup)


async def handle_update_continue(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = update.effective_user.id
    ExcelManager.clear_update_reminder(user_id)
    
    products = ExcelManager.get_user_products(user_id)

    keyboard = [
        [InlineKeyboardButton("➕ Add Product", callback_data="add_product")],
        [InlineKeyboardButton("📋 My Products", callback_data="view_myproducts")],
        [InlineKeyboardButton("📊 Price History", callback_data="view_history")],
        [InlineKeyboardButton("🔙 Back to Menu", callback_data="back_to_menu")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await safe_edit_message(
        query,
        f"✅ <b>Monitoring Continued</b>\n\nYour <b>{len(products)}</b> product(s) will continue to be monitored.\n\nYou'll receive another reminder in {MONTHLY_UPDATE_REMINDER_DAYS} days.",
        reply_markup
    )


async def back_to_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = update.effective_user.id
    country = ExcelManager.get_user_country(user_id) or "US"
    products = ExcelManager.get_user_products(user_id)

    country_flags = {"FR": "🇫🇷", "IT": "🇮🇹", "US": "🇺🇸"}

    message = (
        "🏠 <b>Main Menu</b>\n"
        "━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"🌍 <b>Country:</b> {country_flags.get(country, '')} {country}\n"
        f"📦 <b>Monitored Products:</b> {len(products)}\n\n"
        "Send me an AliExpress link to add a product,\nor use the buttons below:"
    )

    keyboard = [
        [InlineKeyboardButton("➕ Add Product", callback_data="add_product")],
        [InlineKeyboardButton("📋 My Products", callback_data="view_myproducts")],
        [InlineKeyboardButton("📊 Price History", callback_data="view_history")],
        [InlineKeyboardButton("ℹ️ Help", callback_data="show_help")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await safe_edit_message(query, message, reply_markup)


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if query:
        await query.answer()
    
    help_text = (
        "ℹ️ <b>Help - AliExpress Price Monitor</b>\n"
        "━━━━━━━━━━━━━━━━━━━━━\n\n"
        "<b>🔧 How it works:</b>\n"
        "1️⃣ Select your country\n"
        "2️⃣ Send product links to monitor\n"
        "3️⃣ Get notified when prices change!\n\n"
        "<b>📋 Commands:</b>\n"
        "/start - Start the bot\n"
        "/help - Show this help\n"
        "/myproducts - View monitored products\n"
        "/history - View price history\n\n"
        "<b>⚡ Fast Monitoring:</b>\n"
        f"• Checks {CONCURRENT_REQUESTS} products simultaneously\n"
        f"• Updates every {MONITORING_INTERVAL//60} minutes\n"
        "• Instant notifications on price changes\n\n"
        "💡 <i>Optimized for speed and efficiency!</i>"
    )

    keyboard = [
        [InlineKeyboardButton("➕ Add Product", callback_data="add_product")],
        [InlineKeyboardButton("🔙 Back to Menu", callback_data="back_to_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    if query:
        await safe_edit_message(query, help_text, reply_markup)
    else:
        await update.message.reply_text(help_text, reply_markup=reply_markup, parse_mode='HTML')


# ============================================================================
# PRICE MONITORING
# ============================================================================

async def check_single_product(api: AliExpressAPI, product: Dict, context: ContextTypes.DEFAULT_TYPE) -> Dict:
    try:
        user_country = ExcelManager.get_user_country(product['user_id']) or product['country']
        
        result = await api.get_product_details(product['product_id'], user_country)
        
        if not result.get("success"):
            ExcelManager.update_product_price(
                product['user_id'],
                product['product_id'],
                product['current_price'] or 0,
                user_country
            )
            return {'success': False, 'product_id': product['product_id'], 'error': result.get('error')}
        
        new_price = result['price']
        old_price = product['current_price'] or 0
        
        ExcelManager.update_product_price(
            product['user_id'],
            product['product_id'],
            new_price,
            user_country,
            result['product_url']
        )
        
        price_changed = abs(new_price - old_price) > 0.01
        
        if price_changed:
            ExcelManager.save_price_change(
                user_id=product['user_id'],
                product_id=product['product_id'],
                title=product['title'],
                old_price=old_price,
                new_price=new_price,
                currency=product['currency'] or 'USD'
            )
            
            change = new_price - old_price
            change_percent = (change / old_price * 100) if old_price > 0 else 0
            
            emoji = "📉 PRICE DROP!" if change < 0 else "📈 PRICE INCREASE"
            
            affiliate_link = await api.generate_affiliate_link(result['product_url'], user_country)
            
            notification = (
                f"{emoji}\n\n"
                f"<b>{product['title'][:80]}...</b>\n\n"
                f"💵 <b>Old:</b> ${old_price:.2f}\n"
                f"💵 <b>New:</b> ${new_price:.2f}\n"
                f"📊 <b>Change:</b> ${change:+.2f} ({change_percent:+.1f}%)\n"
            )
            
            if change < 0:
                notification += f"💰 <b>You Save:</b> ${abs(change):.2f}\n"
            
            notification += f"\n🔗 <code>{affiliate_link}</code>"
            
            keyboard = [
                [InlineKeyboardButton("🛒 Buy Now", url=affiliate_link)],
                [InlineKeyboardButton("📊 View History", callback_data="view_history")],
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            try:
                if product.get('image_url'):
                    await context.bot.send_photo(
                        chat_id=product['user_id'],
                        photo=product['image_url'],
                        caption=notification,
                        reply_markup=reply_markup,
                        parse_mode='HTML'
                    )
                else:
                    await context.bot.send_message(
                        chat_id=product['user_id'],
                        text=notification,
                        reply_markup=reply_markup,
                        parse_mode='HTML'
                    )
            except Exception as e:
                logger.error(f"Notification failed for user {product['user_id']}: {e}")
        
        return {'success': True, 'product_id': product['product_id'], 'changed': price_changed}
        
    except Exception as e:
        logger.error(f"Error checking {product['product_id']}: {e}")
        return {'success': False, 'product_id': product['product_id'], 'error': str(e)}


async def monitor_prices(context: ContextTypes.DEFAULT_TYPE):
    logger.info(f"🔍 MONITORING CYCLE - {datetime.now().strftime('%H:%M:%S')}")
    
    products_to_check = ExcelManager.get_products_to_check(PRODUCTS_PER_CYCLE)
    
    if not products_to_check:
        logger.info("No products to check")
        return

    logger.info(f"Checking {len(products_to_check)} products...")
    
    api = await get_api_instance()
    
    price_changes = 0
    checked = 0
    errors = 0
    
    for i in range(0, len(products_to_check), CONCURRENT_REQUESTS):
        batch = products_to_check[i:i + CONCURRENT_REQUESTS]
        
        tasks = [check_single_product(api, product, context) for product in batch]
        results = await asyncio.gather(*tasks, return_exceptions=True)
        
        for result in results:
            if isinstance(result, Exception):
                errors += 1
                continue
                
            if result.get('success'):
                checked += 1
                if result.get('changed'):
                    price_changes += 1
            else:
                errors += 1
        
        if i + CONCURRENT_REQUESTS < len(products_to_check):
            await asyncio.sleep(REQUEST_DELAY)
    
    logger.info(f"✅ Checked: {checked}, Changes: {price_changes}, Errors: {errors}")


async def send_monthly_reminder_job(context: ContextTypes.DEFAULT_TYPE):
    user_id = context.job.data
    
    products = ExcelManager.get_user_products(user_id)
    
    if not products:
        return

    ExcelManager.set_update_reminder(user_id)

    message = (
        "🔔 <b>Monthly Product List Update</b>\n"
        "━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"You are currently monitoring <b>{len(products)}</b> product(s).\n\n"
        "Would you like to:\n"
        "• ✅ Continue monitoring current products\n"
        "• 🗑️ Delete some products\n"
        "• ➕ Add new products\n\n"
        f"⚠️ <b>Please respond within {UPDATE_RESPONSE_DEADLINE_DAYS} days</b>"
    )

    keyboard = [
        [InlineKeyboardButton("✅ Continue Monitoring", callback_data="update_continue")],
        [InlineKeyboardButton("➕ Add Products", callback_data="add_product")],
        [InlineKeyboardButton("🗑️ Manage Products", callback_data="manage_products")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    try:
        await context.bot.send_message(
            chat_id=user_id,
            text=message,
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
        logger.info(f"Reminder sent to user {user_id}")
    except Exception as e:
        logger.error(f"Error sending reminder to user {user_id}: {e}")


async def check_monthly_updates(context: ContextTypes.DEFAULT_TYPE):
    logger.info("🔔 Checking monthly updates")
    
    users_need_reminder = ExcelManager.get_users_needing_reminder()
    
    for user_id in users_need_reminder:
        products = ExcelManager.get_user_products(user_id)
        if products:
            context.job_queue.run_once(
                send_monthly_reminder_job,
                when=1,
                data=user_id
            )
            await asyncio.sleep(2)
    
    users_past_deadline = ExcelManager.get_users_past_deadline()
    
    for user_id in users_past_deadline:
        logger.info(f"Cleaning up user {user_id}")
        
        try:
            await context.bot.send_message(
                chat_id=user_id,
                text=(
                    "⚠️ <b>Products Removed</b>\n\n"
                    "Your monitored products have been removed due to no response.\n\n"
                    "You can start monitoring again anytime by using /start"
                ),
                parse_mode='HTML'
            )
        except:
            pass
        
        ExcelManager.delete_all_user_data(user_id)
        ExcelManager.clear_update_reminder(user_id)


# ============================================================================
# ERROR HANDLER
# ============================================================================

async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle errors in the dispatcher."""
    error = context.error
    
    # Handle Conflict error (another instance running)
    if isinstance(error, Conflict):
        logger.error("="*60)
        logger.error("❌ CONFLICT ERROR: Another bot instance is running!")
        logger.error("Please stop all other instances of this bot.")
        logger.error("="*60)
        return
    
    # Handle network errors
    if isinstance(error, (NetworkError, TimedOut)):
        logger.warning(f"Network error: {error}")
        return
    
    # Log other errors
    logger.error(f"Update {update} caused error: {error}")
    logger.error(traceback.format_exc())


async def post_init(application):
    """Called after the application is initialized"""
    logger.info("✅ Bot initialized successfully")


async def post_shutdown(application):
    """Called after the application is shut down"""
    global api_instance
    logger.info("🛑 Shutting down...")
    if api_instance:
        await api_instance.close_session()
    logger.info("👋 Goodbye!")


# ============================================================================
# MAIN
# ============================================================================

def main():
    """Main function to start the bot"""
    logger.info("="*60)
    logger.info("📦 Initializing bot...")
    logger.info("="*60)
    
    # Initialize Excel files
    try:
        ExcelManager.init_excel_files()
        logger.info("✅ Excel files ready")
    except Exception as e:
        logger.error(f"❌ Failed to initialize Excel files: {e}")
        logger.error(traceback.format_exc())
        sys.exit(1)

    # Clear webhook and pending updates BEFORE starting
    logger.info("🔄 Clearing webhook and old updates...")
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    
    try:
        # Verify bot token
        bot_valid = loop.run_until_complete(get_bot_info())
        if not bot_valid:
            logger.error("❌ Invalid bot token! Please check your TELEGRAM_BOT_TOKEN")
            sys.exit(1)
        
        # Clear webhook
        loop.run_until_complete(clear_webhook_and_updates())
        
        # Wait for Telegram to release connections
        logger.info("⏳ Waiting for old connections to close...")
        time.sleep(3)
        
    except Exception as e:
        logger.warning(f"Pre-start cleanup warning: {e}")
    finally:
        loop.close()

    # Build the application
    try:
        logger.info("🔧 Building application...")
        application = (
            Application.builder()
            .token(TELEGRAM_BOT_TOKEN)
            .post_init(post_init)
            .post_shutdown(post_shutdown)
            .build()
        )
        logger.info("✅ Application built")
    except Exception as e:
        logger.error(f"❌ Failed to build application: {e}")
        logger.error(traceback.format_exc())
        sys.exit(1)

    # Add error handler
    application.add_error_handler(error_handler)

    # Create conversation handler
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            SELECTING_COUNTRY: [
                CallbackQueryHandler(country_selected, pattern="^country_")
            ],
            ENTERING_LINK: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_link),
                CallbackQueryHandler(add_product_prompt, pattern="^add_product$"),
                CallbackQueryHandler(view_my_products, pattern="^view_myproducts$"),
                CallbackQueryHandler(view_price_history, pattern="^view_history$"),
                CallbackQueryHandler(back_to_menu, pattern="^back_to_menu$"),
                CallbackQueryHandler(help_command, pattern="^show_help$"),
            ],
        },
        fallbacks=[CommandHandler("start", start)],
        allow_reentry=True,
    )

    # Add handlers
    application.add_handler(conv_handler)
    application.add_handler(CallbackQueryHandler(add_product_prompt, pattern="^add_product$"))
    application.add_handler(CallbackQueryHandler(view_my_products, pattern="^view_myproducts$"))
    application.add_handler(CallbackQueryHandler(manage_products, pattern="^manage_products$"))
    application.add_handler(CallbackQueryHandler(delete_product_callback, pattern="^delete_"))
    application.add_handler(CallbackQueryHandler(view_price_history, pattern="^view_history$"))
    application.add_handler(CallbackQueryHandler(show_price_history, pattern="^history_"))
    application.add_handler(CallbackQueryHandler(handle_update_continue, pattern="^update_continue$"))
    application.add_handler(CallbackQueryHandler(back_to_menu, pattern="^back_to_menu$"))
    application.add_handler(CallbackQueryHandler(help_command, pattern="^show_help$"))
    
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("myproducts", view_my_products))
    application.add_handler(CommandHandler("history", view_price_history))
    
    logger.info("✅ Handlers registered")

    # Setup job queue
    job_queue = application.job_queue
    
    job_queue.run_repeating(
        monitor_prices,
        interval=MONITORING_INTERVAL,
        first=30
    )
    
    job_queue.run_repeating(
        check_monthly_updates,
        interval=MONTHLY_CHECK_INTERVAL,
        first=120
    )
    
    logger.info("✅ Job queue configured")

    logger.info("="*60)
    logger.info("🚀 BOT IS STARTING...")
    logger.info("="*60)

    # Run the bot with drop_pending_updates=True
    try:
        application.run_polling(
            allowed_updates=Update.ALL_TYPES,
            drop_pending_updates=True,
            close_loop=False
        )
    except Conflict:
        logger.error("="*60)
        logger.error("❌ CONFLICT: Another bot instance is already running!")
        logger.error("Please make sure only ONE instance of this bot is running.")
        logger.error("="*60)
        sys.exit(1)
    except Exception as e:
        logger.error(f"❌ Bot crashed: {e}")
        logger.error(traceback.format_exc())
        sys.exit(1)


if __name__ == '__main__':
    main()
