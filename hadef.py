import os
import re
import time
import asyncio
import hashlib
import hmac
import aiohttp
import io
from datetime import datetime, timedelta
from typing import Dict, Any, Optional, List
from urllib.parse import urlparse, parse_qs, urlencode, urlunparse
import openpyxl
from openpyxl import Workbook, load_workbook
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ConversationHandler,
    ContextTypes,
    filters
)
from telegram.error import BadRequest
from flask import Flask
import threading
import boto3
from botocore.exceptions import ClientError

# ============================================================================
# CONFIGURATION
# ============================================================================
TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "8354835888:AAF_F1KR40K6nmI_RwkDPwUa74L__CNuY3s")
ALIEXPRESS_APP_KEY = os.environ.get("ALIEXPRESS_APP_KEY", "519492")
ALIEXPRESS_APP_SECRET = os.environ.get("ALIEXPRESS_APP_SECRET", "R2Zl1pe2p47dFFjXz30546XTwu4JcFlk")
ALIEXPRESS_TRACKING_ID = os.environ.get("ALIEXPRESS_TRACKING_ID", "hadef")

# AWS S3 Configuration
AWS_ACCESS_KEY_ID = os.environ.get("AWS_ACCESS_KEY_ID", "")
AWS_SECRET_ACCESS_KEY = os.environ.get("AWS_SECRET_ACCESS_KEY", "")
AWS_BUCKET_NAME = os.environ.get("AWS_BUCKET_NAME", "telegram-bot-storage")
AWS_REGION = os.environ.get("AWS_REGION", "us-east-1")

# Excel files
USERS_FILE = "users.xlsx"
PRODUCTS_FILE = "products.xlsx"
PRICE_HISTORY_FILE = "price_history.xlsx"

# Monitoring configuration
CONCURRENT_REQUESTS = 10
REQUEST_DELAY = 1
MONITORING_INTERVAL = 300
PRODUCTS_PER_CYCLE = 100
MAX_CHECK_INTERVAL_HOURS = 24

# Rate limit configuration
RATE_LIMIT_RETRY_DELAY = 30
MAX_RETRIES = 3
REQUEST_TIMEOUT = 15

# Monthly update configuration
MONTHLY_UPDATE_REMINDER_DAYS = 30
UPDATE_RESPONSE_DEADLINE_DAYS = 3
MONTHLY_CHECK_INTERVAL = 86400

# States for conversation
SELECTING_COUNTRY, ENTERING_LINK, CHANGING_COUNTRY, MANAGING_PRODUCTS, VIEWING_HISTORY = range(5)

# ============================================================================
# S3 STORAGE MANAGER
# ============================================================================
class S3StorageManager:
    """Manage Excel files in AWS S3 or local storage"""
    
    def __init__(self):
        self.use_s3 = bool(AWS_ACCESS_KEY_ID and AWS_SECRET_ACCESS_KEY)
        if self.use_s3:
            try:
                self.s3_client = boto3.client(
                    's3',
                    aws_access_key_id=AWS_ACCESS_KEY_ID,
                    aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
                    region_name=AWS_REGION
                )
                self.bucket_name = AWS_BUCKET_NAME
                # Test connection
                self.s3_client.head_bucket(Bucket=self.bucket_name)
                print(f"✅ Using AWS S3 storage: {self.bucket_name}")
            except Exception as e:
                print(f"⚠️ S3 connection failed: {e}. Falling back to local storage")
                self.use_s3 = False
        else:
            print("⚠️ Using local storage (not recommended for production)")
    
    def download_file(self, filename: str) -> Optional[Workbook]:
        """Download Excel file from S3 or load locally"""
        try:
            if self.use_s3:
                response = self.s3_client.get_object(Bucket=self.bucket_name, Key=filename)
                file_content = response['Body'].read()
                return load_workbook(io.BytesIO(file_content))
            else:
                if os.path.exists(filename):
                    return load_workbook(filename)
                return None
        except ClientError as e:
            if e.response['Error']['Code'] == 'NoSuchKey':
                return None
            print(f"❌ Error downloading {filename}: {e}")
            return None
        except Exception as e:
            print(f"❌ Error loading {filename}: {e}")
            return None
    
    def upload_file(self, workbook: Workbook, filename: str) -> bool:
        """Upload Excel file to S3 or save locally"""
        try:
            if self.use_s3:
                buffer = io.BytesIO()
                workbook.save(buffer)
                buffer.seek(0)
                
                self.s3_client.put_object(
                    Bucket=self.bucket_name,
                    Key=filename,
                    Body=buffer.getvalue(),
                    ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            else:
                workbook.save(filename)
            return True
        except Exception as e:
            print(f"❌ Error uploading {filename}: {e}")
            return False
    
    def file_exists(self, filename: str) -> bool:
        """Check if file exists in S3 or locally"""
        try:
            if self.use_s3:
                self.s3_client.head_object(Bucket=self.bucket_name, Key=filename)
                return True
            else:
                return os.path.exists(filename)
        except ClientError:
            return False
        except Exception:
            return False

storage = S3StorageManager()

# ============================================================================
# EXCEL MANAGEMENT
# ============================================================================
class ExcelManager:
    @staticmethod
    def init_excel_files():
        """Initialize Excel files if they don't exist"""
        if not storage.file_exists(USERS_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Users"
            ws.append(["User ID", "Username", "Country", "Date Added", "Last Update Reminder", 
                      "Update Deadline", "Needs Update Response"])
            storage.upload_file(wb, USERS_FILE)
            print(f"✅ Created {USERS_FILE}")

        if not storage.file_exists(PRODUCTS_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Products"
            ws.append(["User ID", "Product ID", "Product URL", "Title", "Current Price", 
                      "Original Price", "Currency", "Image URL", "Country", "Date Added", "Last Checked"])
            storage.upload_file(wb, PRODUCTS_FILE)
            print(f"✅ Created {PRODUCTS_FILE}")
        
        if not storage.file_exists(PRICE_HISTORY_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Price History"
            ws.append(["User ID", "Product ID", "Product Title", "Old Price", "New Price", 
                      "Change Amount", "Change Percent", "Currency", "Date"])
            storage.upload_file(wb, PRICE_HISTORY_FILE)
            print(f"✅ Created {PRICE_HISTORY_FILE}")

    @staticmethod
    def save_user(user_id: int, username: str, country: str):
        """Save or update user in Excel"""
        try:
            wb = storage.download_file(USERS_FILE)
            if not wb:
                ExcelManager.init_excel_files()
                wb = storage.download_file(USERS_FILE)
            
            ws = wb.active
            user_exists = False
            for row in ws.iter_rows(min_row=2, max_col=7):
                if row[0].value == user_id:
                    row[2].value = country
                    user_exists = True
                    break
            
            if not user_exists:
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ws.append([user_id, username, country, now, "", "", "No"])
            
            storage.upload_file(wb, USERS_FILE)
        except Exception as e:
            print(f"❌ Error saving user: {e}")

    @staticmethod
    def get_user_country(user_id: int) -> Optional[str]:
        """Get user's current country"""
        try:
            wb = storage.download_file(USERS_FILE)
            if not wb:
                return None
            ws = wb.active
            for row in ws.iter_rows(min_row=2, max_col=3):
                if row[0].value == user_id:
                    return row[2].value
        except Exception as e:
            print(f"❌ Error getting user country: {e}")
        return None

    @staticmethod
    def update_user_products_country(user_id: int, new_country: str):
        """Update country for all products of a user"""
        try:
            wb = storage.download_file(PRODUCTS_FILE)
            if not wb:
                return 0
            ws = wb.active
            updated_count = 0
            for row in ws.iter_rows(min_row=2):
                if row[0].value == user_id:
                    row[8].value = new_country
                    updated_count += 1
            storage.upload_file(wb, PRODUCTS_FILE)
            return updated_count
        except Exception as e:
            print(f"❌ Error updating products country: {e}")
            return 0

    @staticmethod
    def save_product(user_id: int, product_id: str, product_url: str, title: str, 
                    price: float, original_price: float, currency: str, image_url: str, country: str):
        """Save product to Excel"""
        try:
            wb = storage.download_file(PRODUCTS_FILE)
            if not wb:
                ExcelManager.init_excel_files()
                wb = storage.download_file(PRODUCTS_FILE)
            ws = wb.active
            product_exists = False
            for row in ws.iter_rows(min_row=2):
                if row[0].value == user_id and row[1].value == product_id:
                    row[2].value = product_url
                    row[4].value = price
                    row[5].value = original_price
                    row[8].value = country
                    row[10].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    product_exists = True
                    break
            if not product_exists:
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ws.append([user_id, product_id, product_url, title, price, original_price, 
                          currency, image_url, country, now, now])
            storage.upload_file(wb, PRODUCTS_FILE)
        except Exception as e:
            print(f"❌ Error saving product: {e}")

    @staticmethod
    def get_all_products() -> List[Dict]:
        """Get all products for monitoring"""
        products = []
        try:
            wb = storage.download_file(PRODUCTS_FILE)
            if not wb:
                return products
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    products.append({
                        'user_id': row[0], 'product_id': row[1], 'product_url': row[2],
                        'title': row[3], 'current_price': row[4], 'original_price': row[5],
                        'currency': row[6], 'image_url': row[7], 'country': row[8],
                        'date_added': row[9], 'last_checked': row[10]
                    })
        except Exception as e:
            print(f"❌ Error getting products: {e}")
        return products

    @staticmethod
    def get_products_to_check(limit: int) -> List[Dict]:
        """Get products that need to be checked"""
        all_products = ExcelManager.get_all_products()
        if not all_products:
            return []
        def get_last_checked_time(product):
            last_checked = product.get('last_checked')
            if not last_checked or last_checked == 'Never' or last_checked == '':
                return datetime.min
            try:
                return datetime.strptime(str(last_checked), "%Y-%m-%d %H:%M:%S")
            except:
                return datetime.min
        sorted_products = sorted(all_products, key=get_last_checked_time)
        return sorted_products[:limit]

    @staticmethod
    def update_product_price(user_id: int, product_id: str, new_price: float, country: str = None, product_url: str = None):
        """Update product price and last checked time"""
        try:
            wb = storage.download_file(PRODUCTS_FILE)
            if not wb:
                return
            ws = wb.active
            for row in ws.iter_rows(min_row=2):
                if row[0].value == user_id and row[1].value == product_id:
                    if product_url:
                        row[2].value = product_url
                    row[4].value = new_price
                    if country:
                        row[8].value = country
                    row[10].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    break
            storage.upload_file(wb, PRODUCTS_FILE)
        except Exception as e:
            print(f"❌ Error updating product price: {e}")

    @staticmethod
    def save_price_change(user_id: int, product_id: str, title: str, old_price: float, 
                         new_price: float, currency: str):
        """Save price change to history"""
        try:
            if abs(new_price - old_price) < 0.01:
                return
            wb = storage.download_file(PRICE_HISTORY_FILE)
            if not wb:
                ExcelManager.init_excel_files()
                wb = storage.download_file(PRICE_HISTORY_FILE)
            ws = wb.active
            change = new_price - old_price
            change_percent = ((new_price - old_price) / old_price * 100) if old_price > 0 else 0
            ws.append([user_id, product_id, title, old_price, new_price, round(change, 2),
                      round(change_percent, 2), currency, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            storage.upload_file(wb, PRICE_HISTORY_FILE)
            print(f"✅ Price change: {title} - ${change:+.2f} ({change_percent:+.1f}%)")
        except Exception as e:
            print(f"❌ Error saving price change: {e}")

    @staticmethod
    def get_user_products(user_id: int) -> List[Dict]:
        """Get all products for a specific user"""
        all_products = ExcelManager.get_all_products()
        return [p for p in all_products if p['user_id'] == user_id]

    @staticmethod
    def delete_product(user_id: int, product_id: str) -> bool:
        """Delete a product from monitoring"""
        try:
            wb = storage.download_file(PRODUCTS_FILE)
            if not wb:
                return False
            ws = wb.active
            rows_to_delete = []
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row[0].value == user_id and row[1].value == product_id:
                    rows_to_delete.append(idx)
            for row_idx in reversed(rows_to_delete):
                ws.delete_rows(row_idx)
            storage.upload_file(wb, PRODUCTS_FILE)
            return True
        except Exception as e:
            print(f"❌ Error deleting product: {e}")
            return False

    @staticmethod
    def get_price_history(user_id: int, product_id: str, months: int = None) -> List[Dict]:
        """Get price history for a product"""
        history = []
        try:
            wb = storage.download_file(PRICE_HISTORY_FILE)
            if not wb:
                return history
            ws = wb.active
            cutoff_date = None
            if months:
                cutoff_date = datetime.now() - timedelta(days=months * 30)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == user_id and row[1] == product_id:
                    record_date = datetime.strptime(str(row[8]), "%Y-%m-%d %H:%M:%S")
                    if cutoff_date and record_date < cutoff_date:
                        continue
                    history.append({
                        'title': row[2], 'old_price': row[3], 'new_price': row[4],
                        'change_amount': row[5], 'change_percent': row[6],
                        'currency': row[7], 'date': row[8]
                    })
        except Exception as e:
            print(f"❌ Error getting price history: {e}")
        return sorted(history, key=lambda x: x['date'], reverse=True)

# ============================================================================
# ALIEXPRESS API
# ============================================================================
class AliExpressAPI:
    def __init__(self, app_key: str, app_secret: str, tracking_id: str):
        self.app_key = app_key
        self.app_secret = app_secret
        self.tracking_id = tracking_id
        self.api_url = "https://api-sg.aliexpress.com/sync"
        self.timeout = aiohttp.ClientTimeout(total=REQUEST_TIMEOUT)
        self.session = None

    async def get_session(self):
        """Get or create aiohttp session"""
        if self.session is None or self.session.closed:
            connector = aiohttp.TCPConnector(limit=50, limit_per_host=10, ttl_dns_cache=300)
            self.session = aiohttp.ClientSession(connector=connector, timeout=self.timeout)
        return self.session

    async def close_session(self):
        """Close aiohttp session"""
        if self.session and not self.session.closed:
            await self.session.close()

    @staticmethod
    def _now_ms() -> str:
        return str(int(time.time() * 1000))

    def generate_signature(self, params: Dict[str, Any]) -> str:
        params_to_sign = {k: str(v) for k, v in params.items() if k != "sign" and v is not None and v != ""}
        sorted_items = sorted(params_to_sign.items(), key=lambda x: x[0])
        canonical = "".join(f"{k}{v}" for k, v in sorted_items)
        signature = hmac.new(self.app_secret.encode("utf-8"), canonical.encode("utf-8"), hashlib.md5).hexdigest().upper()
        return signature

    @staticmethod
    def extract_product_id(url: str) -> Optional[str]:
        patterns = [r"/item/(\d+)\.html", r"/i/(\d+)\.html", r"/(\d+)\.html", r"item/(\d+)", 
                   r"/goods/(\d+)", r"product/(\d+)", r"/dp/(\d+)"]
        for p in patterns:
            m = re.search(p, url)
            if m:
                return m.group(1)
        return None

    @staticmethod
    def build_product_url(product_id: str) -> str:
        return f"https://www.aliexpress.com/item/{product_id}.html"

    async def resolve_shortened_url(self, url: str, max_retries: int = 3) -> str:
        """Async URL resolver"""
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
        session = await self.get_session()
        for attempt in range(max_retries):
            try:
                async with session.head(url, allow_redirects=True, headers=headers) as response:
                    final_url = str(response.url)
                    if self.extract_product_id(final_url):
                        return final_url
            except Exception:
                if attempt < max_retries - 1:
                    await asyncio.sleep(2)
                    continue
                return url
        return url

    @staticmethod
    def is_shortened_url(url: str) -> bool:
        patterns = ["s.click.aliexpress.com", "a.aliexpress.com", "/e/_", "ali.ski"]
        return any(pattern in url.lower() for pattern in patterns)

    @staticmethod
    def is_rate_limited(error_msg: str) -> bool:
        patterns = ["frequency exceeds the limit", "rate limit", "too many requests"]
        return any(pattern in error_msg.lower() for pattern in patterns)

    async def get_product_details(self, product_id: str, country: str = "US", retry_count: int = 0) -> Dict[str, Any]:
        """Async product details fetcher"""
        start_time = time.time()
        method = "aliexpress.affiliate.productdetail.get"
        params = {
            "app_key": self.app_key, "format": "json", "method": method, "sign_method": "hmac",
            "timestamp": self._now_ms(), "v": "2.0", "tracking_id": self.tracking_id,
            "product_ids": str(product_id), "target_currency": "USD", "target_language": "EN", "country": country
        }
        params["sign"] = self.generate_signature(params)
        session = await self.get_session()
        try:
            async with session.get(self.api_url, params=params) as response:
                response.raise_for_status()
                data = await response.json()
                elapsed_time = time.time() - start_time
                if "error_response" in data:
                    error_msg = data["error_response"].get("msg", "API Error")
                    if self.is_rate_limited(error_msg) and retry_count < MAX_RETRIES:
                        wait_time = RATE_LIMIT_RETRY_DELAY * (2 ** retry_count)
                        await asyncio.sleep(wait_time)
                        return await self.get_product_details(product_id, country, retry_count + 1)
                    return {"success": False, "error": error_msg, "time_taken": elapsed_time}
                resp_key = None
                for k in data.keys():
                    if k.endswith("_response"):
                        resp_key = k
                        break
                if not resp_key:
                    return {"success": False, "error": "Invalid response", "time_taken": elapsed_time}
                resp_data = data[resp_key]
                result = resp_data.get("resp_result", {}).get("result", {})
                products = result.get("products", {}).get("product", [])
                if not products:
                    return {"success": False, "error": "Product not found", "time_taken": elapsed_time}
                product = products[0] if isinstance(products, list) else products
                sale_price = product.get("target_sale_price") or product.get("sale_price")
                original_price = product.get("target_original_price") or product.get("original_price")
                def to_float(val):
                    if val is None:
                        return None
                    try:
                        price_str = str(val).replace("USD", "").replace("$", "").replace(",", "").strip()
                        return float(price_str)
                    except:
                        return None
                current_price = to_float(sale_price)
                orig_price = to_float(original_price)
                if current_price is None:
                    return {"success": False, "error": "No price available", "time_taken": elapsed_time}
                product_url = self.build_product_url(str(product.get("product_id", product_id)))
                return {
                    "success": True, "product_id": str(product.get("product_id", product_id)),
                    "title": product.get("product_title", "N/A"), "price": current_price,
                    "original_price": orig_price or current_price, "currency": "USD",
                    "image_url": product.get("product_main_image_url", ""), "product_url": product_url,
                    "time_taken": elapsed_time
                }
        except asyncio.TimeoutError:
            elapsed_time = time.time() - start_time
            return {"success": False, "error": "Request timeout", "time_taken": elapsed_time}
        except Exception as e:
            elapsed_time = time.time() - start_time
            return {"success": False, "error": str(e), "time_taken": elapsed_time}

    async def generate_affiliate_link(self, product_url: str, country: str = "US") -> Optional[str]:
        """Async affiliate link generator"""
        method = "aliexpress.affiliate.link.generate"
        params = {
            "app_key": self.app_key, "format": "json", "method": method, "sign_method": "hmac",
            "timestamp": self._now_ms(), "v": "2.0", "tracking_id": self.tracking_id,
            "promotion_link_type": "0", "source_values": product_url
        }
        params["sign"] = self.generate_signature(params)
        session = await self.get_session()
        try:
            async with session.post(self.api_url, data=params) as response:
                data = await response.json()
                if "error_response" in data:
                    return product_url
                resp_data = data.get("aliexpress_affiliate_link_generate_response", {})
                result = resp_data.get("resp_result", {})
                if result.get("resp_code") == 200:
                    links = result.get("result", {}).get("promotion_links", {}).get("promotion_link", [])
                    if links:
                        return links[0].get("promotion_link")
                return product_url
        except Exception:
            return product_url

api_instance = None

async def get_api_instance():
    """Get or create global API instance"""
    global api_instance
    if api_instance is None:
        api_instance = AliExpressAPI(ALIEXPRESS_APP_KEY, ALIEXPRESS_APP_SECRET, ALIEXPRESS_TRACKING_ID)
    return api_instance

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================
async def safe_edit_message(query, text, reply_markup=None, parse_mode='HTML'):
    """Safely edit message or send new one if editing fails"""
    try:
        await query.edit_message_text(text, reply_markup=reply_markup, parse_mode=parse_mode)
    except BadRequest:
        try:
            await query.message.delete()
        except:
            pass
        await query.message.reply_text(text, reply_markup=reply_markup, parse_mode=parse_mode)
    except Exception as e:
        print(f"Error editing message: {e}")
        try:
            await query.message.reply_text(text, reply_markup=reply_markup, parse_mode=parse_mode)
        except:
            pass

# ============================================================================
# TELEGRAM BOT HANDLERS
# ============================================================================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /start command"""
    user = update.effective_user
    user_id = user.id
    username = user.username or user.first_name
    keyboard = [
        [InlineKeyboardButton("🇫🇷 France", callback_data="country_FR"),
         InlineKeyboardButton("🇮🇹 Italy", callback_data="country_IT")],
        [InlineKeyboardButton("🇺🇸 United States", callback_data="country_US")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    welcome_text = (
        f"👋 <b>Welcome {username}!</b>\n\n🛍️ <b>AliExpress Price Monitor Bot</b>\n"
        "━━━━━━━━━━━━━━━━━━━━━\n\nI will help you track AliExpress product prices and notify you of changes!\n\n"
        "📍 <b>Please select your country:</b>"
    )
    await update.message.reply_text(welcome_text, reply_markup=reply_markup, parse_mode='HTML')
    return SELECTING_COUNTRY

async def country_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle country selection"""
    query = update.callback_query
    await query.answer()
    country = query.data.split("_")[1]
    user = update.effective_user
    user_id = user.id
    username = user.username or user.first_name
    ExcelManager.save_user(user_id, username, country)
    updated_count = ExcelManager.update_user_products_country(user_id, country)
    context.user_data['country'] = country
    country_flags = {"FR": "🇫🇷", "IT": "🇮🇹", "US": "🇺🇸"}
    message = f"✅ <b>Country Selected: {country_flags.get(country, '')} {country}</b>\n\n"
    if updated_count > 0:
        message += f"🔄 Updated {updated_count} existing products\n\n"
    message += (
        "📎 <b>Now send me an AliExpress product link:</b>\n\n<i>Supported formats:</i>\n"
        "• <code>https://www.aliexpress.com/item/xxxxx.html</code>\n"
        "• <code>https://s.click.aliexpress.com/e/_xxxxx</code>\n\n💡 Or use the menu below:"
    )
    keyboard = [
        [InlineKeyboardButton("➕ Add Product", callback_data="add_product")],
        [InlineKeyboardButton("📋 My Products", callback_data="view_myproducts")],
        [InlineKeyboardButton("📊 Price History", callback_data="view_history")],
        [InlineKeyboardButton("ℹ️ Help", callback_data="show_help")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await safe_edit_message(query, message, reply_markup)
    return ENTERING_LINK

async def add_product_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show prompt to add a product"""
    query = update.callback_query
    await query.answer()
    message = (
        "📎 <b>Send me an AliExpress product link:</b>\n\n<i>Supported formats:</i>\n"
        "• <code>https://www.aliexpress.com/item/xxxxx.html</code>\n"
        "• <code>https://s.click.aliexpress.com/e/_xxxxx</code>\n\n💡 Just paste the link and send it to me!"
    )
    keyboard = [
        [InlineKeyboardButton("📋 My Products", callback_data="view_myproducts")],
        [InlineKeyboardButton("🔙 Back to Menu", callback_data="back_to_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await safe_edit_message(query, message, reply_markup)
    return ENTERING_LINK

async def handle_link(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle product link submission"""
    if not update.message or not update.message.text:
        return ENTERING_LINK
    total_start_time = time.time()
    user_id = update.effective_user.id
    product_url = update.message.text.strip()
    if "aliexpress" not in product_url.lower():
        keyboard = [
            [InlineKeyboardButton("➕ Add Product", callback_data="add_product")],
            [InlineKeyboardButton("📋 My Products", callback_data="view_myproducts")],
            [InlineKeyboardButton("🔙 Back to Menu", callback_data="back_to_menu")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(
            "❌ <b>Invalid Link</b>\n\nPlease send a valid AliExpress product link.",
            reply_markup=reply_markup, parse_mode='HTML'
        )
        return ENTERING_LINK
    processing_msg = await update.message.reply_text("⏳ Processing...")
    country = ExcelManager.get_user_country(user_id) or "US"
    api = await get_api_instance()
    url_start = time.time()
    if api.is_shortened_url(product_url):
        await processing_msg.edit_text("🔗 Resolving shortened URL...")
        product_url = await api.resolve_shortened_url(product_url)
    url_time = time.time() - url_start
    product_id = api.extract_product_id(product_url)
    if not product_id:
        keyboard = [
            [InlineKeyboardButton("➕ Try Again", callback_data="add_product")],
            [InlineKeyboardButton("🔙 Back to Menu", callback_data="back_to_menu")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await processing_msg.edit_text(
            "❌ <b>Could not extract product ID</b>\n\nPlease send a valid AliExpress link.",
            reply_markup=reply_markup, parse_mode='HTML'
        )
        return ENTERING_LINK
    await processing_msg.edit_text("📊 Fetching product details...")
    result = await api.get_product_details(product_id, country)
    api_time = result.get('time_taken', 0)
    if not result.get("success"):
        keyboard = [
            [InlineKeyboardButton("➕ Try Another", callback_data="add_product")],
            [InlineKeyboardButton("🔙 Back to Menu", callback_data="back_to_menu")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await processing_msg.edit_text(
            f"❌ <b>Cannot monitor this product</b>\n\n<b>Reason:</b> {result.get('error')}\n"
            f"<b>Time taken:</b> {api_time:.2f}s\n\n💡 Try another product or check back later.",
            reply_markup=reply_markup, parse_mode='HTML'
        )
        return ENTERING_LINK
    affiliate_start = time.time()
    affiliate_link = await api.generate_affiliate_link(result['product_url'], country)
    affiliate_time = time.time() - affiliate_start
    save_start = time.time()
    ExcelManager.save_product(
        user_id=user_id, product_id=product_id, product_url=affiliate_link, title=result['title'],
        price=result['price'], original_price=result['original_price'], currency=result['currency'],
        image_url=result['image_url'], country=country
    )
    save_time = time.time() - save_start
    total_time = time.time() - total_start_time
    discount = result['original_price'] - result['price']
    discount_percent = (discount / result['original_price'] * 100) if result['original_price'] > 0 else 0
    message = (
        "✅ <b>Product Added Successfully!</b>\n━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"📦 <b>{result['title'][:80]}...</b>\n\n💵 <b>Current:</b> ${result['price']:.2f}\n"
        f"💰 <b>Original:</b> ${result['original_price']:.2f}\n"
    )
    if discount > 0:
        message += f"🏷️ <b>Discount:</b> ${discount:.2f} ({discount_percent:.1f}% OFF)\n"
    message += (
        f"🌍 <b>Country:</b> {country}\n🆔 <b>ID:</b> {product_id}\n\n"
        f"⏱️ <b>Processing Time:</b>\n   • API call: {api_time:.2f}s\n   • Total: {total_time:.2f}s\n\n"
        f"🔔 <b>Price monitoring active!</b>\n\n🔗 <code>{affiliate_link}</code>"
    )
    keyboard = [
        [InlineKeyboardButton("➕ Add Another Product", callback_data="add_product")],
        [InlineKeyboardButton("📋 My Products", callback_data="view_myproducts")],
        [InlineKeyboardButton("📊 Price History", callback_data="view_history")],
        [InlineKeyboardButton("🔙 Back to Menu", callback_data="back_to_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    if result.get('image_url'):
        try:
            await update.message.reply_photo(
                photo=result['image_url'], caption=message, reply_markup=reply_markup, parse_mode='HTML'
            )
            await processing_msg.delete()
        except:
            await processing_msg.edit_text(message, reply_markup=reply_markup, parse_mode='HTML')
    else:
        await processing_msg.edit_text(message, reply_markup=reply_markup, parse_mode='HTML')
    print(f"   ⏱️ Product {product_id} added - API: {api_time:.2f}s, Total: {total_time:.2f}s")
    return ENTERING_LINK

async def view_my_products(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show user's products"""
    query = update.callback_query
    if query:
        await query.answer()
    user_id = update.effective_user.id
    products = ExcelManager.get_user_products(user_id)
    if not products:
        keyboard = [
            [InlineKeyboardButton("➕ Add Product", callback_data="add_product")],
            [InlineKeyboardButton("🔙 Back", callback_data="back_to_menu")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        message = (
            "📭 <b>No Products Yet</b>\n\nYou haven't added any products to monitor.\n\n"
            "Click 'Add Product' to start tracking prices!"
        )
        if query:
            await safe_edit_message(query, message, reply_markup)
        else:
            await update.message.reply_text(message, reply_markup=reply_markup, parse_mode='HTML')
        return MANAGING_PRODUCTS
    message = f"📋 <b>Your Monitored Products ({len(products)})</b>\n━━━━━━━━━━━━━━━━━━━━━\n\n"
    for idx, product in enumerate(products[:10], 1):
        title = product['title'][:40] + "..." if len(product['title']) > 40 else product['title']
        price = product['current_price']
        original_price = product['original_price']
        discount = original_price - price
        discount_percent = (discount / original_price * 100) if original_price > 0 else 0
        message += f"<b>{idx}. {title}</b>\n"
        message += f"💵 Current: ${price:.2f}"
        if discount > 0:
            message += f" ({discount_percent:.0f}% OFF)"
        message += f"\n🆔 ID: <code>{product['product_id']}</code>\n\n"
    if len(products) > 10:
        message += f"<i>... and {len(products) - 10} more products</i>\n\n"
    message += "💡 <b>Select a product to manage:</b>"
    keyboard = []
    for product in products[:10]:
        title = product['title'][:30] + "..." if len(product['title']) > 30 else product['title']
        keyboard.append([InlineKeyboardButton(
            f"📦 {title}", callback_data=f"manage_{product['product_id']}"
        )])
    keyboard.append([InlineKeyboardButton("➕ Add Product", callback_data="add_product")])
    keyboard.append([InlineKeyboardButton("🔙 Back to Menu", callback_data="back_to_menu")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    if query:
        await safe_edit_message(query, message, reply_markup)
    else:
        await update.message.reply_text(message, reply_markup=reply_markup, parse_mode='HTML')
    return MANAGING_PRODUCTS

async def manage_product(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show product management options"""
    query = update.callback_query
    await query.answer()
    product_id = query.data.split("_")[1]
    user_id = update.effective_user.id
    products = ExcelManager.get_user_products(user_id)
    product = next((p for p in products if p['product_id'] == product_id), None)
    if not product:
        await query.edit_message_text("❌ Product not found")
        return MANAGING_PRODUCTS
    title = product['title']
    price = product['current_price']
    original_price = product['original_price']
    discount = original_price - price
    discount_percent = (discount / original_price * 100) if original_price > 0 else 0
    last_checked = product.get('last_checked', 'Never')
    message = (
        f"📦 <b>{title[:60]}...</b>\n━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"💵 <b>Current Price:</b> ${price:.2f}\n"
        f"💰 <b>Original Price:</b> ${original_price:.2f}\n"
    )
    if discount > 0:
        message += f"🏷️ <b>Discount:</b> ${discount:.2f} ({discount_percent:.1f}% OFF)\n"
    message += (
        f"🌍 <b>Country:</b> {product['country']}\n"
        f"🆔 <b>Product ID:</b> <code>{product_id}</code>\n"
        f"🕐 <b>Last Checked:</b> {last_checked}\n\n"
        f"🔗 <code>{product['product_url']}</code>\n\n"
        f"<b>What would you like to do?</b>"
    )
    keyboard = [
        [InlineKeyboardButton("🔄 Check Price Now", callback_data=f"check_{product_id}")],
        [InlineKeyboardButton("📊 View History", callback_data=f"history_{product_id}")],
        [InlineKeyboardButton("🗑️ Delete Product", callback_data=f"delete_{product_id}")],
        [InlineKeyboardButton("📋 Back to Products", callback_data="view_myproducts")],
        [InlineKeyboardButton("🔙 Main Menu", callback_data="back_to_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    if product.get('image_url'):
        try:
            await query.message.delete()
            await query.message.reply_photo(
                photo=product['image_url'], caption=message, reply_markup=reply_markup, parse_mode='HTML'
            )
            return MANAGING_PRODUCTS
        except:
            pass
    await safe_edit_message(query, message, reply_markup)
    return MANAGING_PRODUCTS

async def check_price_now(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Check product price immediately"""
    query = update.callback_query
    await query.answer("Checking price...")
    product_id = query.data.split("_")[1]
    user_id = update.effective_user.id
    products = ExcelManager.get_user_products(user_id)
    product = next((p for p in products if p['product_id'] == product_id), None)
    if not product:
        await query.edit_message_text("❌ Product not found")
        return MANAGING_PRODUCTS
    await query.edit_message_text("⏳ Checking current price...")
    api = await get_api_instance()
    result = await api.get_product_details(product_id, product['country'])
    if not result.get('success'):
        keyboard = [[InlineKeyboardButton("🔙 Back", callback_data=f"manage_{product_id}")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(
            f"❌ <b>Failed to check price</b>\n\n<b>Error:</b> {result.get('error')}",
            reply_markup=reply_markup, parse_mode='HTML'
        )
        return MANAGING_PRODUCTS
    new_price = result['price']
    old_price = product['current_price']
    price_change = new_price - old_price
    change_percent = (price_change / old_price * 100) if old_price > 0 else 0
    if abs(price_change) > 0.01:
        ExcelManager.save_price_change(user_id, product_id, product['title'], old_price, new_price, product['currency'])
    ExcelManager.update_product_price(user_id, product_id, new_price, product['country'])
    message = (
        f"✅ <b>Price Check Complete</b>\n━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"📦 <b>{product['title'][:60]}...</b>\n\n"
        f"💵 <b>Current Price:</b> ${new_price:.2f}\n"
        f"📊 <b>Previous Price:</b> ${old_price:.2f}\n"
    )
    if abs(price_change) > 0.01:
        if price_change < 0:
            message += f"📉 <b>Price Drop:</b> ${abs(price_change):.2f} ({abs(change_percent):.1f}% OFF) 🎉\n"
        else:
            message += f"📈 <b>Price Increase:</b> ${price_change:.2f} ({change_percent:.1f}%) ⚠️\n"
    else:
        message += "➖ <b>No Price Change</b>\n"
    message += f"\n🕐 <b>Checked at:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    keyboard = [
        [InlineKeyboardButton("📊 View History", callback_data=f"history_{product_id}")],
        [InlineKeyboardButton("🔙 Back to Product", callback_data=f"manage_{product_id}")],
        [InlineKeyboardButton("📋 My Products", callback_data="view_myproducts")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(message, reply_markup=reply_markup, parse_mode='HTML')
    return MANAGING_PRODUCTS

async def view_product_history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """View price history for a product"""
    query = update.callback_query
    await query.answer()
    product_id = query.data.split("_")[1]
    user_id = update.effective_user.id
    products = ExcelManager.get_user_products(user_id)
    product = next((p for p in products if p['product_id'] == product_id), None)
    if not product:
        await query.edit_message_text("❌ Product not found")
        return VIEWING_HISTORY
    history = ExcelManager.get_price_history(user_id, product_id, months=3)
    if not history:
        keyboard = [[InlineKeyboardButton("🔙 Back", callback_data=f"manage_{product_id}")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await safe_edit_message(
            query,
            f"📊 <b>Price History</b>\n━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"📦 <b>{product['title'][:60]}...</b>\n\n"
            f"No price changes recorded yet.\n\nPrice changes will appear here once monitoring detects them.",
            reply_markup
        )
        return VIEWING_HISTORY
    message = (
        f"📊 <b>Price History (Last 3 months)</b>\n━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"📦 <b>{product['title'][:60]}...</b>\n\n"
    )
    for idx, record in enumerate(history[:10], 1):
        change = record['change_amount']
        percent = record['change_percent']
        date = record['date']
        if change < 0:
            icon = "📉"
            change_text = f"${abs(change):.2f} ({abs(percent):.1f}% OFF)"
        else:
            icon = "📈"
            change_text = f"${change:.2f} (+{percent:.1f}%)"
        message += (
            f"{icon} <b>{date}</b>\n"
            f"   ${record['old_price']:.2f} → ${record['new_price']:.2f}\n"
            f"   Change: {change_text}\n\n"
        )
    if len(history) > 10:
        message += f"<i>... and {len(history) - 10} more changes</i>"
    keyboard = [
        [InlineKeyboardButton("🔄 Check Price Now", callback_data=f"check_{product_id}")],
        [InlineKeyboardButton("🔙 Back to Product", callback_data=f"manage_{product_id}")],
        [InlineKeyboardButton("📋 My Products", callback_data="view_myproducts")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await safe_edit_message(query, message, reply_markup)
    return VIEWING_HISTORY

async def delete_product(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Delete a product from monitoring"""
    query = update.callback_query
    await query.answer()
    product_id = query.data.split("_")[1]
    user_id = update.effective_user.id
    products = ExcelManager.get_user_products(user_id)
    product = next((p for p in products if p['product_id'] == product_id), None)
    if not product:
        await query.edit_message_text("❌ Product not found")
        return MANAGING_PRODUCTS
    message = (
        f"⚠️ <b>Confirm Deletion</b>\n━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"📦 <b>{product['title'][:60]}...</b>\n\n"
        f"Are you sure you want to stop monitoring this product?\n\n"
        f"<i>Price history will be preserved.</i>"
    )
    keyboard = [
        [InlineKeyboardButton("✅ Yes, Delete", callback_data=f"confirm_delete_{product_id}"),
         InlineKeyboardButton("❌ Cancel", callback_data=f"manage_{product_id}")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await safe_edit_message(query, message, reply_markup)
    return MANAGING_PRODUCTS

async def confirm_delete_product(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Confirm product deletion"""
    query = update.callback_query
    await query.answer()
    product_id = query.data.split("_")[2]
    user_id = update.effective_user.id
    success = ExcelManager.delete_product(user_id, product_id)
    if success:
        message = "✅ <b>Product Deleted</b>\n\nThe product has been removed from monitoring."
    else:
        message = "❌ <b>Deletion Failed</b>\n\nCould not delete the product. Please try again."
    keyboard = [
        [InlineKeyboardButton("📋 My Products", callback_data="view_myproducts")],
        [InlineKeyboardButton("🔙 Main Menu", callback_data="back_to_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await safe_edit_message(query, message, reply_markup)
    return MANAGING_PRODUCTS

async def show_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show help information"""
    query = update.callback_query
    await query.answer()
    message = (
        "ℹ️ <b>Help & Information</b>\n━━━━━━━━━━━━━━━━━━━━━\n\n"
        "<b>🤖 What I Do:</b>\n"
        "• Monitor AliExpress product prices\n"
        "• Send notifications when prices change\n"
        "• Track price history\n"
        "• Generate affiliate links\n\n"
        "<b>📝 How to Use:</b>\n"
        "1. Select your country\n"
        "2. Send me product links\n"
        "3. I'll monitor prices automatically\n"
        "4. Get notified of price changes\n\n"
        "<b>🔗 Supported Links:</b>\n"
        "• Regular: aliexpress.com/item/...\n"
        "• Shortened: s.click.aliexpress.com/...\n"
        "• App links: a.aliexpress.com/...\n\n"
        "<b>⏰ Monitoring:</b>\n"
        "• Checks every 5 minutes\n"
        "• Notifies you of price changes\n"
        "• Keeps 3 months of history\n\n"
        "<b>💡 Tips:</b>\n"
        "• Check price manually anytime\n"
        "• View history to track trends\n"
        "• Delete products you're not interested in"
    )
    keyboard = [
        [InlineKeyboardButton("➕ Add Product", callback_data="add_product")],
        [InlineKeyboardButton("📋 My Products", callback_data="view_myproducts")],
        [InlineKeyboardButton("🔙 Main Menu", callback_data="back_to_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await safe_edit_message(query, message, reply_markup)
    return ENTERING_LINK

async def back_to_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Return to main menu"""
    query = update.callback_query
    await query.answer()
    user_id = update.effective_user.id
    country = ExcelManager.get_user_country(user_id) or "US"
    country_flags = {"FR": "🇫🇷", "IT": "🇮🇹", "US": "🇺🇸"}
    message = (
        f"🏠 <b>Main Menu</b>\n━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"📍 <b>Your Country:</b> {country_flags.get(country, '')} {country}\n\n"
        f"<b>What would you like to do?</b>"
    )
    keyboard = [
        [InlineKeyboardButton("➕ Add Product", callback_data="add_product")],
        [InlineKeyboardButton("📋 My Products", callback_data="view_myproducts")],
        [InlineKeyboardButton("📊 Price History", callback_data="view_all_history")],
        [InlineKeyboardButton("🌍 Change Country", callback_data="change_country")],
        [InlineKeyboardButton("ℹ️ Help", callback_data="show_help")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await safe_edit_message(query, message, reply_markup)
    return ENTERING_LINK

async def change_country(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Change user's country"""
    query = update.callback_query
    await query.answer()
    keyboard = [
        [InlineKeyboardButton("🇫🇷 France", callback_data="country_FR"),
         InlineKeyboardButton("🇮🇹 Italy", callback_data="country_IT")],
        [InlineKeyboardButton("🇺🇸 United States", callback_data="country_US")],
        [InlineKeyboardButton("🔙 Back", callback_data="back_to_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    message = (
        "🌍 <b>Change Country</b>\n━━━━━━━━━━━━━━━━━━━━━\n\n"
        "Select your new country:\n\n"
        "<i>This will update all your monitored products.</i>"
    )
    await safe_edit_message(query, message, reply_markup)
    return CHANGING_COUNTRY

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Cancel the conversation"""
    await update.message.reply_text("Operation cancelled. Use /start to begin again.")
    return ConversationHandler.END

# ============================================================================
# PRICE MONITORING BACKGROUND TASK
# ============================================================================
async def monitor_prices(application):
    """Background task to monitor product prices"""
    print("🔍 Price monitoring started")
    while True:
        try:
            await asyncio.sleep(MONITORING_INTERVAL)
            products = ExcelManager.get_products_to_check(PRODUCTS_PER_CYCLE)
            if not products:
                print("   No products to check")
                continue
            print(f"   Checking {len(products)} products...")
            api = await get_api_instance()
            for i in range(0, len(products), CONCURRENT_REQUESTS):
                batch = products[i:i + CONCURRENT_REQUESTS]
                tasks = []
                for product in batch:
                    task = api.get_product_details(product['product_id'], product['country'])
                    tasks.append((product, task))
                results = await asyncio.gather(*[task for _, task in tasks])
                for (product, _), result in zip(tasks, results):
                    if result.get('success'):
                        new_price = result['price']
                        old_price = product['current_price']
                        price_change = new_price - old_price
                        if abs(price_change) > 0.01:
                            ExcelManager.save_price_change(
                                product['user_id'], product['product_id'], product['title'],
                                old_price, new_price, product['currency']
                            )
                            try:
                                change_percent = (price_change / old_price * 100) if old_price > 0 else 0
                                if price_change < 0:
                                    emoji = "📉"
                                    message = (
                                        f"{emoji} <b>Price Drop Alert!</b>\n\n"
                                        f"📦 <b>{product['title'][:60]}...</b>\n\n"
                                        f"💵 New Price: ${new_price:.2f}\n"
                                        f"📊 Old Price: ${old_price:.2f}\n"
                                        f"🎉 Saved: ${abs(price_change):.2f} ({abs(change_percent):.1f}% OFF)\n\n"
                                        f"🔗 <code>{product['product_url']}</code>"
                                    )
                                else:
                                    emoji = "📈"
                                    message = (
                                        f"{emoji} <b>Price Increase Alert</b>\n\n"
                                        f"📦 <b>{product['title'][:60]}...</b>\n\n"
                                        f"💵 New Price: ${new_price:.2f}\n"
                                        f"📊 Old Price: ${old_price:.2f}\n"
                                        f"⚠️ Increase: ${price_change:.2f} (+{change_percent:.1f}%)\n\n"
                                        f"🔗 <code>{product['product_url']}</code>"
                                    )
                                keyboard = [
                                    [InlineKeyboardButton("📦 View Product", callback_data=f"manage_{product['product_id']}")],
                                    [InlineKeyboardButton("📋 My Products", callback_data="view_myproducts")]
                                ]
                                reply_markup = InlineKeyboardMarkup(keyboard)
                                await application.bot.send_message(
                                    chat_id=product['user_id'],
                                    text=message,
                                    reply_markup=reply_markup,
                                    parse_mode='HTML'
                                )
                            except Exception as e:
                                print(f"   ❌ Failed to send notification: {e}")
                        ExcelManager.update_product_price(
                            product['user_id'], product['product_id'], new_price, product['country']
                        )
                await asyncio.sleep(REQUEST_DELAY)
        except Exception as e:
            print(f"❌ Error in price monitoring: {e}")

# ============================================================================
# FLASK WEB SERVER
# ============================================================================
app = Flask(__name__)

@app.route('/')
def home():
    return "✅ Telegram Bot is running!"

@app.route('/health')
def health():
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}

def run_flask():
    """Run Flask server"""
    port = int(os.environ.get("PORT", 8000))
    app.run(host='0.0.0.0', port=port, debug=False)

# ============================================================================
# MAIN FUNCTION
# ============================================================================
async def main():
    """Main function to run the bot"""
    print("🚀 Starting Telegram Bot...")
    ExcelManager.init_excel_files()
    application = Application.builder().token(TELEGRAM_BOT_TOKEN).build()
    
    # Conversation handler
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            SELECTING_COUNTRY: [CallbackQueryHandler(country_selected, pattern="^country_")],
            ENTERING_LINK: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_link),
                CallbackQueryHandler(add_product_prompt, pattern="^add_product$"),
                CallbackQueryHandler(view_my_products, pattern="^view_myproducts$"),
                CallbackQueryHandler(show_help, pattern="^show_help$"),
                CallbackQueryHandler(back_to_menu, pattern="^back_to_menu$"),
                CallbackQueryHandler(change_country, pattern="^change_country$")
            ],
            CHANGING_COUNTRY: [CallbackQueryHandler(country_selected, pattern="^country_")],
            MANAGING_PRODUCTS: [
                CallbackQueryHandler(view_my_products, pattern="^view_myproducts$"),
                CallbackQueryHandler(manage_product, pattern="^manage_"),
                CallbackQueryHandler(check_price_now, pattern="^check_"),
                CallbackQueryHandler(view_product_history, pattern="^history_"),
                CallbackQueryHandler(delete_product, pattern="^delete_"),
                CallbackQueryHandler(confirm_delete_product, pattern="^confirm_delete_"),
                CallbackQueryHandler(add_product_prompt, pattern="^add_product$"),
                CallbackQueryHandler(back_to_menu, pattern="^back_to_menu$")
            ],
            VIEWING_HISTORY: [
                CallbackQueryHandler(view_product_history, pattern="^history_"),
                CallbackQueryHandler(manage_product, pattern="^manage_"),
                CallbackQueryHandler(check_price_now, pattern="^check_"),
                CallbackQueryHandler(view_my_products, pattern="^view_myproducts$"),
                CallbackQueryHandler(back_to_menu, pattern="^back_to_menu$")
            ]
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    application.add_handler(conv_handler)
    
    # Start Flask server
    flask_thread = threading.Thread(target=run_flask, daemon=True)
    flask_thread.start()
    print(f"✅ Flask server started on port {os.environ.get('PORT', 8000)}")
    
    # Initialize and start bot
    await application.initialize()
    await application.start()
    await application.updater.start_polling(drop_pending_updates=True)
    print("✅ Bot is running!")
    
    # Start price monitoring
    asyncio.create_task(monitor_prices(application))
    
    # Keep running
    try:
        while True:
            await asyncio.sleep(1)
    except KeyboardInterrupt:
        print("\n⚠️ Stopping bot...")
        await application.stop()

if __name__ == "__main__":
    asyncio.run(main())
